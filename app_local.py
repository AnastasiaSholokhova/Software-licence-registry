from flask import Flask, request, session, redirect, url_for, render_template, flash, Response, send_file
import psycopg2
import psycopg2.extras
import re 
from werkzeug.security import generate_password_hash, check_password_hash
import io
from functools import wraps
import xlwt
from datetime import datetime, date, timedelta
from waitress import serve
from flask_mail import Mail
from random import randint
import smtplib
from email.mime.text import MIMEText
import os
import openpyxl
import pandas as pd

# инициализация приложения
app = Flask(__name__)
app.secret_key = os.urandom(24)
mail = Mail(app)

#переменные для базы данных
DBNAME = "Users" #База данных пользователей
DBNAME2 = "software_licences" #База данных лицензирования ПО
DBUSER = "postgres" #Имя пользователя для сервера postgresql
DBPASS = "12345" #Пароль от сервера (pgadmin4)

# Данные для отправки писем на адрес электронной почты - например, для смены пароля
app.config['MAIL_SERVER'] = 'smtp.yandex.ru' #Сервер для электронной почты
app.config['MAIL_PORT'] = 587 #Порт для электронной почты
app.config['MAIL_USERNAME'] = '' #Адрес электронной почты
app.config['MAIL_PASSWORD'] = '' #Пароль от электронной почты
app.config['MAIL_USE_TLS'] = 1
app.config['MAIL_USE_SSL'] = False

#Установка соединения с базами данных
conn = psycopg2.connect(dbname=DBNAME, user=DBUSER, password=DBPASS)
conn2 = psycopg2.connect(dbname=DBNAME2, user=DBUSER, password=DBPASS)


@app.route('/register', methods=['GET', 'POST'])
def register():
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
 
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form and 'email' in request.form and 'phone_number' in request.form:
        fullname = request.form['fullname']
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        phone_number = request.form['phone_number']
        role = request.form['role']
    
        _hashed_password = generate_password_hash(password)
        cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
        account = cursor.fetchone()
        cursor.execute('SELECT * FROM users WHERE email=%s', (email,))
        email_exists = cursor.fetchone()
        if account:
            flash('Аккаунт уже зарегистрирован!')
        if email_exists:
            flash('Этот адрес электронной почты уже используется!')
        elif not re.match(r'[^@]+@[^@]+\.[^@]+', email):
            flash('Некорректный адрес электронной почты!')
        elif not re.match(r'[A-Za-z0-9]+', username):
            flash('Имя пользователя должно содержать только буквы и цифры!')
        elif not re.match(r"^(?:\+?(?:[0-9] ?){6,14}[0-9])$", phone_number):
            flash('Некорректный номер телефона!')
        elif not username or not password or not email or not phone_number or not role:
            flash('Пожалуйста, заполните форму!')
        else:
            cursor.execute("INSERT INTO users (fullname, username, password, email, phone_number, role) VALUES (%s,%s,%s,%s,%s,%s)", (fullname, username, _hashed_password, email, phone_number, role))
            conn.commit()
            cursor.execute('SELECT LASTVAL() AS id;')
            user_id = cursor.fetchone()
            session['loggedin'] = True
            session['id'] = user_id
            session['username'] = username
            session['role'] = role
            flash('Вы были успешно зарегистрированы!')
            return redirect(url_for('login'))
    elif request.method == 'POST':
        flash('Пожалуйста, заполните форму!')
    return render_template('register.html')

@app.route('/', methods=['GET', 'POST'])
def login():
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
   
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form:
        username = request.form['username']
        password = request.form['password']
        
        cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
        account = cursor.fetchone()
 
        if account:
            password_rs = account['password']
            if check_password_hash(password_rs, password):
                selected_role = account['role']
                session['loggedin'] = True
                session['id'] = account['id']
                session['username'] = account['username']
                session['role'] = selected_role
                session['email'] = account['email']
                if selected_role == 'editor':
                    return redirect(url_for('home_editor'))
                elif selected_role == 'support':
                    return redirect(url_for('home_support'))
                else:
                    return redirect(url_for('home'))
            else:
                flash('Неверное имя пользователя/пароль')
        else:
            flash('Неверное имя пользователя/пароль')
 
    return render_template('login.html')

@app.route('/login_ad', methods=['GET', 'POST'])
def login_ad():
    pass

def role_required(*roles):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if 'role' in session and session['role'] in roles:
                return func(*args, **kwargs)
            else:
                flash('У вас нет доступа к этой странице!')
                return redirect(url_for('login'))
        return wrapper
    return decorator

@app.route('/logout')
def logout():
   session.pop('loggedin', None)
   session.pop('id', None)
   session.pop('username', None)
   return redirect(url_for('login'))

@app.route('/total')
@role_required('admin')
def total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права FROM общая_информация'
        cur.execute(s)
        list_total = cur.fetchall()
        return render_template('total.html', username=session['username'], list_total = list_total)
    return redirect(url_for('login'))

@app.route('/editor_total')
@role_required('editor')
def editor_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права FROM общая_информация'
        cur.execute(s)
        list_total = cur.fetchall()
        return render_template('editor_total.html', username=session['username'], list_total=list_total)
    return redirect(url_for('login'))

@app.route('/show_total', methods=['POST', 'GET'])
@role_required('admin')
def show_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права FROM общая_информация'
        cur.execute(s)
        list_total = cur.fetchall()
        return render_template('add_total.html', username=session['username'], list_total = list_total)
    return redirect(url_for('login'))

@app.route('/editor_show_total', methods=['POST', 'GET'])
@role_required('editor')
def editor_show_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права FROM общая_информация'
        cur.execute(s)
        list_total = cur.fetchall()
        return render_template('editor_show_total.html', username=session['username'], list_total=list_total)
    return redirect(url_for('login'))

@app.route('/support_show_total')
@role_required('support')
def support_show_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT * FROM общая_информация'
        cur.execute(s)
        list_total = cur.fetchall()
        return render_template('support_show_total.html', username=session['username'], list_total=list_total)
    return redirect(url_for('login'))

@app.route('/add_total', methods=['POST'])
def add_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if request.method == 'POST':
            статья_затрат = request.form['статья_затрат']
            наименование_ПО_БУ = request.form['наименование_ПО_БУ']
            наименование_ПО = request.form['наименование_ПО']
            краткое_наименование_ПО = request.form['краткое_наименование_ПО']
            количество = request.form['количество']
            код = request.form['код']
            филиал = request.form['филиал']
            счет_затрат = request.form['счет_затрат']
            вид_деятельности = request.form['вид_деятельности']
            стоимость_ПО_без_ндс = request.form['стоимость_ПО_без_ндс']
            стоимость_ПО_с_ндс = request.form['стоимость_ПО_с_ндс']
            срок_полезного_использования_мес = request.form['срок_полезного_использования_мес']
            дата_начала_списания = request.form['дата_начала_списания']
            дата_окончания_списания = request.form['дата_окончания_списания']
            договор_счет = request.form['договор_счет']
            контрагент = request.form['контрагент']
            первичный_документ = request.form['первичный_документ']
            страна_производитель = request.form['страна_производитель']
            правообладатель = request.form['правообладатель']
            включен_в_реестр = request.form['включен_в_реестр']
            срок_предоставления_права = request.form['срок_предоставления_права']
            cur.execute(
                "INSERT INTO общая_информация (статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (
                    статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности,
                    стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания,
                    дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель,
                    правообладатель, включен_в_реестр, срок_предоставления_права
                )
            )
            #cur.execute('INSERT INTO Контрагенты (наименование_контрагента, договор) VALUES(%s, %s) WHERE ', (наименование_контрагента, договор, ))
            conn2.commit()
            flash('Запись была успешно добавлена!')
            return redirect(url_for('total'))
        return redirect(url_for('total'))

@app.route('/editor_add_total', methods=['POST'])
@role_required('editor')
def editor_add_total():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if request.method == 'POST':
            статья_затрат = request.form['статья_затрат']
            наименование_ПО_БУ = request.form['наименование_ПО_БУ']
            наименование_ПО = request.form['наименование_ПО']
            краткое_наименование_ПО = request.form['краткое_наименование_ПО']
            количество = request.form['количество']
            код = request.form['код']
            филиал = request.form['филиал']
            счет_затрат = request.form['счет_затрат']
            вид_деятельности = request.form['вид_деятельности']
            стоимость_ПО_без_ндс = request.form['стоимость_ПО_без_ндс']
            стоимость_ПО_с_ндс = request.form['стоимость_ПО_с_ндс']
            срок_полезного_использования_мес = request.form['срок_полезного_использования_мес']
            дата_начала_списания = request.form['дата_начала_списания']
            дата_окончания_списания = request.form['дата_окончания_списания']
            договор_счет = request.form['договор_счет']
            контрагент = request.form['контрагент']
            первичный_документ = request.form['первичный_документ']
            страна_производитель = request.form['страна_производитель']
            правообладатель = request.form['правообладатель']
            включен_в_реестр = request.form['включен_в_реестр']
            срок_предоставления_права = request.form['срок_предоставления_права']
            cur.execute(
                "INSERT INTO общая_информация (статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (
                    статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности,
                    стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес, дата_начала_списания,
                    дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель,
                    правообладатель, включен_в_реестр, срок_предоставления_права
                )
            )
            #cur.execute('INSERT INTO Контрагенты (наименование_контрагента, договор) VALUES(%s, %s) WHERE ', (наименование_контрагента, договор, ))
            conn2.commit()
            flash('Запись была успешно добавлена!')
            return redirect(url_for('editor_total'))
        return redirect(url_for('editor_total'))

@app.route('/upload_total', methods=['POST', 'GET'])
def upload_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=6):
                    data.append(row)
                for row in data:
                    дата_начала_списания = datetime.strptime(row[10], '%d.%m.%Y').date()
                    дата_окончания_списания = datetime.strptime(row[11], '%d.%m.%Y').date()
                    cur.execute('''
    INSERT INTO общая_информация (
        статья_затрат,
        наименование_ПО_БУ,
        количество,
        код,
        филиал,
        счет_затрат,
        вид_деятельности,
        стоимость_ПО_без_ндс,
        стоимость_ПО_с_ндс,
        срок_полезного_использования_мес,
        дата_начала_списания,
        дата_окончания_списания,
        договор_счет,
        контрагент,
        первичный_документ,
        страна_производитель,
        правообладатель,
        включен_в_реестр,
        срок_предоставления_права
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', 
    (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], 
    дата_начала_списания, дата_окончания_списания, row[12], row[13], row[14], row[15], row[16], row[17], row[18]))
                    cur.execute('''INSERT INTO лицензии (
                            наименование_ПО, контрагент,
                            дата_начала_списания, дата_окончания_списания,
                            счёт_списания, стоимость_ПО,
                            количество, итоговая_стоимость_ПО,
                            страна_производитель, код) 
                            VALUES (%s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s) RETURNING номер''', 
                        (row[1], row[13], дата_начала_списания,
                         дата_окончания_списания, row[9],
                         row[7], row[2], row[8], row[15], row[3]))
    
                    licence_id = cur.fetchone()[0]
                    общая_стоимость = row[7]
    
                    software_sign = 'Новое' if общая_стоимость == row[8] else 'Техподдержка'
    
                        # Обновление лицензии
                    cur.execute('''UPDATE лицензии SET признак_ПО = %s,
                                          срок_действия_лицензии = GREATEST(to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW(), interval '0 seconds')
                                          WHERE номер = %s''', 
                                    (software_sign, дата_окончания_списания, licence_id))
                    cur.execute('SELECT COUNT(*) FROM Контрагенты WHERE наименование_контрагента=%s', (row[13],))
                    count = cur.fetchone()[0]
                    if count == 0:
                        cur.execute("""INSERT INTO Контрагенты (наименование_контрагента, договор)
                                    VALUES (%s, %s)""", (row[13], row[12]))
                    cur.execute('INSERT INTO Виды_ПО (наименование_ПО, вид_ПО) VALUES(%s, %s)', (row[1], row[1]))
                    conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('total'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('total'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('total'))
        return render_template('total.html')
    else:
        return render_template('total.html')
    
@app.route('/editor_upload_total', methods=['POST', 'GET'])
def editor_upload_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=6):
                    data.append(row)
                for row in data:
                    дата_начала_списания = datetime.strptime(row[10], '%d.%m.%Y').date()
                    дата_окончания_списания = datetime.strptime(row[11], '%d.%m.%Y').date()
                    cur.execute('''
    INSERT INTO общая_информация (
        статья_затрат,
        наименование_ПО_БУ,
        количество,
        код,
        филиал,
        счет_затрат,
        вид_деятельности,
        стоимость_ПО_без_ндс,
        стоимость_ПО_с_ндс,
        срок_полезного_использования_мес,
        дата_начала_списания,
        дата_окончания_списания,
        договор_счет,
        контрагент,
        первичный_документ,
        страна_производитель,
        правообладатель,
        включен_в_реестр,
        срок_предоставления_права
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', 
    (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], 
    дата_начала_списания, дата_окончания_списания, row[12], row[13], row[14], row[15], row[16], row[17], row[18]))
                    cur.execute('''INSERT INTO лицензии (
                            наименование_ПО, контрагент,
                            дата_начала_списания, дата_окончания_списания,
                            счёт_списания, стоимость_ПО,
                            количество, итоговая_стоимость_ПО,
                            страна_производитель, код) 
                            VALUES (%s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s) RETURNING номер''', 
                        (row[1], row[13], дата_начала_списания,
                         дата_окончания_списания, row[9],
                         row[7], row[2], row[8], row[15], row[3]))
    
                    licence_id = cur.fetchone()[0]
                    общая_стоимость = row[7]
    
                    software_sign = 'Новое' if общая_стоимость == row[8] else 'Техподдержка'
    
                        # Обновление лицензии
                    cur.execute('''UPDATE лицензии SET признак_ПО = %s,
                                          срок_действия_лицензии = GREATEST(to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW(), interval '0 seconds')
                                          WHERE номер = %s''', 
                                    (software_sign, дата_окончания_списания, licence_id))
                    cur.execute('SELECT COUNT(*) FROM Контрагенты WHERE наименование_контрагента=%s', (row[13],))
                    count = cur.fetchone()[0]
                    if count == 0:
                        cur.execute("""INSERT INTO Контрагенты (наименование_контрагента, договор)
                                    VALUES (%s, %s)""", (row[13], row[12]))
                    cur.execute('INSERT INTO Виды_ПО (наименование_ПО, вид_ПО) VALUES(%s, %s)', (row[1], row[1]))
                    conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_total'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_total'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_total'))
        return render_template('editor_total.html')
    else:
        return render_template('editor_total.html')

@app.route('/download_total_list')
def download_total_list(): 
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM общая_информация')
    res_total = cur.fetchall()
    output_total = io.BytesIO()
    workbook_total = xlwt.Workbook()
    sh_total = workbook_total.add_sheet('Отчет по общей информации')
    sh_total.write(0, 0, 'Статья затрат')
    sh_total.write(0, 1, 'Наименование ПО (БУ)')
    sh_total.write(0, 2, 'Наименование ПО')
    sh_total.write(0, 3, 'Краткое наименование ПО')
    sh_total.write(0, 4, 'Количество')
    sh_total.write(0, 5, 'Код')
    sh_total.write(0, 6, 'Филиал')
    sh_total.write(0, 7, 'Счет затрат')
    sh_total.write(0, 8, 'Вид деятельности')
    sh_total.write(0, 9, 'Стоимость ПО (всего) без НДС')
    sh_total.write(0, 10, 'Стоимость ПО (всего) с НДС')
    sh_total.write(0, 11, 'Срок полезного использования (мес.)')
    sh_total.write(0, 12, 'Дата начала списания')
    sh_total.write(0, 13, 'Дата окончания списания')
    sh_total.write(0, 14, 'Договор/счет')
    sh_total.write(0, 15, 'Контрагент')
    sh_total.write(0, 16, 'Первичный документ')
    sh_total.write(0, 17, 'Страна-производитель')
    sh_total.write(0, 18, 'Правообладатель')
    sh_total.write(0, 19, 'Включено в реестр Российского ПО Да/нет')
    sh_total.write(0, 20, 'Срок предоставления права')
    idx = 0 
    for row in res_total:
        sh_total.write(idx+1, 0, row['статья_затрат'])
        sh_total.write(idx+1, 1, row['наименование_ПО_БУ'])
        sh_total.write(idx+1, 2, row['наименование_ПО'])
        sh_total.write(idx+1, 3, row['краткое_наименование_ПО'])
        sh_total.write(idx+1, 4, row['количество'])
        sh_total.write(idx+1, 5, row['код'])
        sh_total.write(idx+1, 6, row['филиал'])
        sh_total.write(idx+1, 7, row['счет_затрат'])
        sh_total.write(idx+1, 8, row['вид_деятельности'])
        sh_total.write(idx+1, 9, row['стоимость_ПО_без_ндс'])
        sh_total.write(idx+1, 10, row['стоимость_ПО_с_ндс'])
        sh_total.write(idx+1, 11, row['срок_полезного_использования_мес'])
        sh_total.write(idx+1, 12, row['дата_начала_списания'])
        sh_total.write(idx+1, 13, row['дата_окончания_списания'])
        sh_total.write(idx+1, 14, row['договор_счет'])
        sh_total.write(idx+1, 15, row['контрагент'])
        sh_total.write(idx+1, 16, row['первичный_документ'])
        sh_total.write(idx+1, 17, row['страна_производитель'])
        sh_total.write(idx+1, 18, row['правообладатель'])
        sh_total.write(idx+1, 19, row['включен_в_реестр'])
        sh_total.write(idx+1, 20, row['срок_предоставления_права'])
        idx += 1
    workbook_total.save(output_total)
    output_total.seek(0)
    return Response(output_total, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=total_list_report.xls"})
   
@app.route('/edit_total/<id>', methods=['POST', 'GET'])
def get_total(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM общая_информация WHERE код = %s', (id,))
    data = cur.fetchall()
    cur.close()
    return render_template('edit_total.html', total = data)

@app.route('/editor_edit_total/<id>', methods=['POST', 'GET'])
def editor_get_total(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM общая_информация WHERE код = %s', (id, ))
    data = cur.fetchall()
    cur.close()
    return render_template('editor_edit_total.html', total=data)

@app.route('/update_total/<id>', methods=['POST'])
def update_total(id):
    if request.method == 'POST':
        статья_затрат = request.form['статья_затрат']
        наименование_ПО_БУ = request.form['наименование_ПО_БУ']
        наименование_ПО = request.form['наименование_ПО']
        краткое_наименование_ПО = request.form['краткое_наименование_ПО']
        количество = request.form['количество']
        код = request.form['код']
        филиал = request.form['филиал']
        счет_затрат = request.form['счет_затрат']
        вид_деятельности = request.form['вид_деятельности']
        стоимость_ПО_без_ндс = request.form['стоимость_ПО_без_ндс']
        стоимость_ПО_с_ндс = request.form['стоимость_ПО_с_ндс']
        срок_полезного_использования_мес = request.form['срок_полезного_использования_мес']
        дата_начала_списания = request.form['дата_начала_списания']
        дата_окончания_списания = request.form['дата_окончания_списания']
        договор_счет = request.form['договор_счет']
        контрагент = request.form['контрагент']
        первичный_документ = request.form['первичный_документ']
        страна_производитель = request.form['страна_производитель']
        правообладатель = request.form['правообладатель']
        включен_в_реестр = request.form['включен_в_реестр']
        срок_предоставления_права = request.form['срок_предоставления_права']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE общая_информация
                    SET
                    статья_затрат=%s,
                    наименование_ПО_БУ=%s,
                    наименование_ПО=%s,
                    краткое_наименование_ПО=%s,
                    количество=%s,
                    код=%s,
                    филиал=%s,
                    счет_затрат=%s,
                    вид_деятельности=%s,
                    стоимость_ПО_без_ндс=%s,
                    стоимость_ПО_с_ндс=%s,
                    срок_полезного_использования_мес=%s,
                    дата_начала_списания=%s,
                    дата_окончания_списания=%s,
                    договор_счет=%s,
                    контрагент=%s,
                    первичный_документ=%s,
                    страна_производитель=%s,
                    правообладатель=%s,
                    включен_в_реестр=%s,
                    срок_предоставления_права=%s
                    WHERE код=%s
                    """, (статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес,
                    дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('total'))
    
@app.route('/editor_update_total/<id>', methods=['POST'])
def editor_update_total(id):
    if request.method == 'POST':
        статья_затрат = request.form['статья_затрат']
        наименование_ПО_БУ = request.form['наименование_ПО_БУ']
        наименование_ПО = request.form['наименование_ПО']
        краткое_наименование_ПО = request.form['краткое_наименование_ПО']
        количество = request.form['количество']
        код = request.form['код']
        филиал = request.form['филиал']
        счет_затрат = request.form['счет_затрат']
        вид_деятельности = request.form['вид_деятельности']
        стоимость_ПО_без_ндс = request.form['стоимость_ПО_без_ндс']
        стоимость_ПО_с_ндс = request.form['стоимость_ПО_с_ндс']
        срок_полезного_использования_мес = request.form['срок_полезного_использования_мес']
        дата_начала_списания = request.form['дата_начала_списания']
        дата_окончания_списания = request.form['дата_окончания_списания']
        договор_счет = request.form['договор_счет']
        контрагент = request.form['контрагент']
        первичный_документ = request.form['первичный_документ']
        страна_производитель = request.form['страна_производитель']
        правообладатель = request.form['правообладатель']
        включен_в_реестр = request.form['включен_в_реестр']
        срок_предоставления_права = request.form['срок_предоставления_права']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE общая_информация
                    SET
                    статья_затрат=%s,
                    наименование_ПО_БУ=%s,
                    наименование_ПО=%s,
                    краткое_наименование_ПО=%s,
                    количество=%s,
                    код=%s,
                    филиал=%s,
                    счет_затрат=%s,
                    вид_деятельности=%s,
                    стоимость_ПО_без_ндс=%s,
                    стоимость_ПО_с_ндс=%s,
                    срок_полезного_использования_мес=%s,
                    дата_начала_списания=%s,
                    дата_окончания_списания=%s,
                    договор_счет=%s,
                    контрагент=%s,
                    первичный_документ=%s,
                    страна_производитель=%s,
                    правообладатель=%s,
                    включен_в_реестр=%s,
                    срок_предоставления_права=%s
                    WHERE код=%s
                    """, (статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, количество, код, филиал, счет_затрат, вид_деятельности, стоимость_ПО_без_ндс, стоимость_ПО_с_ндс, срок_полезного_использования_мес,
                    дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, включен_в_реестр, срок_предоставления_права, id))
        conn2.commit()
        flash("Запись успешно обновлена!")
        return redirect(url_for('editor_total'))

@app.route('/delete_total', methods=['POST', 'GET'])
def delete_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('delete_total_checkbox'):
            cur.execute('DELETE FROM общая_информация WHERE код = %s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('total'))
    
@app.route('/editor_delete_total', methods=['POST', 'GET'])
def editor_delete_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_delete_total_checkbox'):
            cur.execute('DELETE FROM общая_информация WHERE код = %s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_total'))

@app.route('/delete_all_total', methods=['POST', 'GET'])
def delete_all_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM общая_информация')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('total'))
    
@app.route('/editor_delete_all_total', methods=['POST', 'GET'])
def editor_delete_all_total():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM общая_информация')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('editor_total'))
    
@app.route('/home')
@role_required('admin')
def home():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание, код FROM лицензии'
        cur.execute(s)
        list_licences = cur.fetchall()
        return render_template('home.html', username=session['username'], list_licences = list_licences)
    return redirect(url_for('login'))

@app.route('/home_support')
@role_required('support')
def home_support():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание, код FROM лицензии'
        cur.execute(s)
        list_licences = cur.fetchall()
        return render_template('home_support.html', username=session['username'], list_licences = list_licences)
    return redirect(url_for('login'))

@app.route('/home_editor')
@role_required('editor')
def home_editor():
    if 'loggedin' in session:
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = 'SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание, код FROM лицензии'
        cur.execute(s)
        list_licences = cur.fetchall()
        return render_template('home_editor.html', username=session['username'], list_licences = list_licences)
    return redirect(url_for('login'))

@app.route('/show_licence', methods=['GET', 'POST'])
def show_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM лицензии')
    data = cur.fetchall()
    return render_template('add_licence.html', data=data)

@app.route('/add_licence', methods=['POST'])
def add_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        контрагент = request.form['контрагент']
        дата_начала_списания = request.form['дата_начала_списания']
        дата_окончания_списания = request.form['дата_окончания_списания']
        счёт_списания = request.form['счёт_списания']
        итоговая_стоимость_ПО = request.form['итоговая_стоимость_ПО']
        количество = request.form['количество']
        страна_производитель = request.form['страна_производитель']
        оплачено = request.form['оплачено']
        примечание = request.form['примечание']
        код = request.form['код']

        # Проверка существования ПО
        cur.execute('SELECT 1 FROM Справочник_ПО WHERE наименование_ПО=%s', (наименование_ПО,))
        lic_exists = cur.fetchone()
        
        if lic_exists:
            cur.execute('SELECT стоимость_за_единицу FROM Справочник_ПО WHERE наименование_ПО=%s', (наименование_ПО,))
            inst = cur.fetchone()
            стоимость_ПО = inst['стоимость_за_единицу']

            if not количество:
                flash('Пожалуйста, введите количество ПО!')
                return render_template('add_licence.html')

            if not дата_начала_списания or not дата_окончания_списания:
                flash('Пожалуйста, заполните все поля формы!')
                return render_template('add_licence.html')

            if итоговая_стоимость_ПО == стоимость_ПО * количество:
                признак_ПО = 'Новое'
            else:
                признак_ПО = 'Техподдержка'

            # Рассчет остатка
            if оплачено:
                остаток = 0
            elif счёт_списания == '12' and not оплачено:
                остаток = итоговая_стоимость_ПО
            elif счёт_списания == '36' and not оплачено:
                окончание_действия_лицензии_date = datetime.strptime(дата_окончания_списания, '%Y-%m-%d').date()
                срок_действия_мес = (окончание_действия_лицензии_date - date.today()).days / 30
                мес_стоимость = итоговая_стоимость_ПО / 36
                остаток = round(мес_стоимость * срок_действия_мес, 2)
            else:
                остаток = 0

            # Рассчет срока действия лицензии
            окончание_действия_лицензии_date = datetime.strptime(дата_окончания_списания, '%Y-%m-%d').date()
            срок_действия_лицензии_date = окончание_действия_лицензии_date - datetime.now().date()
            срок_действия_лицензии = str(срок_действия_лицензии_date.days) + ' days' if срок_действия_лицензии_date.days > 0 else '0 days'
            
            # Вставка в базу данных
            cur.execute(
                "INSERT INTO лицензии (наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание, код) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::interval, %s, %s, %s, %s)",
                (
                    наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания,
                    счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО,
                    признак_ПО, страна_производитель, срок_действия_лицензии,
                    оплачено, остаток, примечание, код
                )
            )
            conn2.commit()
            flash('Запись была успешно создана!')
            return redirect(url_for('home'))
        else:
            flash('Лицензия не найдена в справочнике ПО!')
            return render_template('add_licence.html')

    return redirect(url_for('home'))
    
@app.route('/editor_show_licence', methods=['GET', 'POST'])
def editor_show_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM  лицензии')
    data = cur.fetchall()
    return render_template('editor_add_licence.html', data=data)

@app.route('/editor_add_licence', methods=['POST'])
def editor_add_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        контрагент = request.form['контрагент']
        дата_начала_списания = request.form['дата_начала_списания']
        дата_окончания_списания = request.form['дата_окончания_списания']
        счёт_списания = request.form['счёт_списания']
        итоговая_стоимость_ПО = request.form['итоговая_стоимость_ПО']
        количество = request.form['количество']
        страна_производитель = request.form['страна_производитель']
        оплачено = request.form['оплачено']
        примечание = request.form['примечание']
        код = request.form['код']

        # Проверка существования ПО
        cur.execute('SELECT 1 FROM Справочник_ПО WHERE наименование_ПО=%s', (наименование_ПО,))
        lic_exists = cur.fetchone()
        
        if lic_exists:
            cur.execute('SELECT стоимость_за_единицу FROM Справочник_ПО WHERE наименование_ПО=%s', (наименование_ПО,))
            inst = cur.fetchone()
            стоимость_ПО = inst['стоимость_за_единицу']

            if not количество:
                flash('Пожалуйста, введите количество ПО!')
                return render_template('editor_add_licence.html')

            if not дата_начала_списания or not дата_окончания_списания:
                flash('Пожалуйста, заполните все поля формы!')
                return render_template('editor_add_licence.html')

            if итоговая_стоимость_ПО == стоимость_ПО * количество:
                признак_ПО = 'Новое'
            else:
                признак_ПО = 'Техподдержка'

            # Рассчет остатка
            if оплачено:
                остаток = 0
            elif счёт_списания == '12' and not оплачено:
                остаток = итоговая_стоимость_ПО
            elif счёт_списания == '36' and not оплачено:
                окончание_действия_лицензии_date = datetime.strptime(дата_окончания_списания, '%Y-%m-%d').date()
                срок_действия_мес = (окончание_действия_лицензии_date - date.today()).days / 30
                мес_стоимость = итоговая_стоимость_ПО / 36
                остаток = round(мес_стоимость * срок_действия_мес, 2)
            else:
                остаток = 0

            # Рассчет срока действия лицензии
            окончание_действия_лицензии_date = datetime.strptime(дата_окончания_списания, '%Y-%m-%d').date()
            срок_действия_лицензии_date = окончание_действия_лицензии_date - datetime.now().date()
            срок_действия_лицензии = str(срок_действия_лицензии_date.days) + ' days' if срок_действия_лицензии_date.days > 0 else '0 days'
            
            # Вставка в базу данных
            cur.execute(
                "INSERT INTO лицензии (наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание, код) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::interval, %s, %s, %s, %s)",
                (
                    наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания,
                    счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО,
                    признак_ПО, страна_производитель, срок_действия_лицензии,
                    оплачено, остаток, примечание, код
                )
            )
            conn2.commit()
            flash('Запись была успешно создана!')
            return redirect(url_for('home_editor'))
        else:
            flash('Лицензия не найдена в справочнике ПО!')
            return render_template('editor_add_licence.html')

    return redirect(url_for('home_editor'))
    
@app.route('/edit/<id>', methods=['POST', 'GET'])
def get_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание FROM лицензии WHERE код = %s', (id,))
    data = cur.fetchone()
    cur.close()
    return render_template('edit.html', licence=data)

@app.route('/editor_edit/<id>', methods=['POST', 'GET'])
def editor_get_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, счёт_списания, стоимость_ПО, количество, итоговая_стоимость_ПО, признак_ПО, страна_производитель, срок_действия_лицензии, оплачено, остаток, примечание FROM лицензии WHERE код =%s', (id,))
    data = cur.fetchone()
    cur.close()
    return render_template('editor_edit.html', licence=data)

@app.route('/update/<id>', methods=['POST'])
def update_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        вендор = request.form['вендор']
        начало_действия_лицензии = request.form['начало_действия_лицензии']
        счёт_списания = request.form['счёт_списания']
        заказчик_ПО = request.form['заказчик_ПО']
        признак_ПО = request.form['признак_ПО']
        количество_ПО = int(request.form['количество_ПО'])
        оплачено = bool(request.form.get('оплачено'))
        примечание = request.form['примечание']
        cur.execute('SELECT 1 FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s;', (наименование_ПО, признак_ПО))
        lic_exists = cur.fetchone()
        if lic_exists:
            cur.execute('SELECT стоимость_за_единицу FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s;', 
                        (наименование_ПО, признак_ПО,))
            lic = cur.fetchone()
            стоимость_за_единицу = lic['стоимость_за_единицу']
            if not количество_ПО:
                flash('Пожалуйста, введите количество ПО!')
                return redirect(url_for('edit'))
            if признак_ПО == 'Новое':
                итоговая_стоимость = стоимость_за_единицу * количество_ПО
            elif признак_ПО == 'Техподдержка':
                итоговая_стоимость = стоимость_за_единицу * количество_ПО * 1.2
            else:
                итоговая_стоимость = стоимость_за_единицу * количество_ПО
            if счёт_списания == '12':
                начало_действия_лицензии_date = datetime.strptime(начало_действия_лицензии, '%Y-%m-%d').date()
                окончание_действия_лицензии = (начало_действия_лицензии_date + timedelta(days=365)).strftime('%Y-%m-%d')
            elif счёт_списания == '36':
                начало_действия_лицензии_date = datetime.strptime(начало_действия_лицензии, '%Y-%m-%d').date()
                окончание_действия_лицензии = (начало_действия_лицензии_date + timedelta(days=365 * 3)).strftime('%Y-%m-%d')
            if оплачено:
                остаток = 0
            elif счёт_списания == '12' and not оплачено:
                остаток = итоговая_стоимость
            elif счёт_списания == '36' and not оплачено:
                окончание_действия_лицензии_date = datetime.strptime(окончание_действия_лицензии, '%Y-%m-%d').date()
                срок_действия_мес = (окончание_действия_лицензии_date - date.today()).days / 30
                мес_стоимость = итоговая_стоимость / 36
                остаток = round(мес_стоимость * срок_действия_мес, 2)
            else:
                остаток = 0
            cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute(""" UPDATE лицензии
                        SET
                        наименование_ПО=%s,
                        вендор=%s,
                        начало_действия_лицензии=%s,
                        окончание_действия_лицензии=%s,
                        счёт_списания=%s,
                        стоимость_за_единицу=%s,
                        итоговая_стоимость=%s,
                        заказчик_ПО=%s,
                        признак_ПО=%s,
                        количество_ПО=%s,
                        срок_действия_лицензии=GREATEST(to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW(), interval '0 seconds'),
                        оплачено=%s,
                        остаток=%s,
                        примечание=%s
                        WHERE код=%s
                        """, (наименование_ПО, вендор, начало_действия_лицензии, окончание_действия_лицензии, счёт_списания, стоимость_за_единицу, итоговая_стоимость, заказчик_ПО, признак_ПО, количество_ПО, окончание_действия_лицензии, оплачено, остаток, примечание, id))
            flash("Запись успешно обновлена!")
            conn2.commit()
            return redirect(url_for('home'))
        else:
            flash('Такого ПО нет в справочнике ПО!')
            return redirect(url_for('get_licence', id=id))
    
@app.route('/editor_update/<id>', methods=['POST'])
def editor_update_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        вендор = request.form['вендор']
        начало_действия_лицензии = request.form['начало_действия_лицензии']
        счёт_списания = request.form['счёт_списания']
        заказчик_ПО = request.form['заказчик_ПО']
        признак_ПО = request.form['признак_ПО']
        количество_ПО = int(request.form['количество_ПО'])
        оплачено = bool(request.form.get('оплачено'))
        примечание = request.form['примечание']
        cur.execute('SELECT 1 FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s;', (наименование_ПО, признак_ПО,))
        lic_exists = cur.fetchone()
        if lic_exists:
            cur.execute('SELECT стоимость_за_единицу FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s;',
                        (наименование_ПО, признак_ПО,))
            lic = cur.fetchone()
            стоимость_за_единицу = lic['стоимость_за_единицу']
            if not количество_ПО:
                flash('Пожалуйста, введите количество ПО!')
                return redirect(url_for('editor_edit'))
            if признак_ПО == 'Новое':
                итоговая_стоимость = стоимость_за_единицу * количество_ПО
            elif признак_ПО == 'Техподдержка':
                итоговая_стоимость = стоимость_за_единицу * количество_ПО * 1.2
            else:
                итоговая_стоимость = стоимость_за_единицу * количество_ПО
            if счёт_списания == '12':
                начало_действия_лицензии_date = datetime.strptime(начало_действия_лицензии, '%Y-%m-%d').date()
                окончание_действия_лицензии = (начало_действия_лицензии_date + timedelta(days=365)).strftime('%Y-%m-%d')
            elif счёт_списания == '36':
                начало_действия_лицензии_date = datetime.strptime(начало_действия_лицензии, '%Y-%m-%d').date()
                окончание_действия_лицензии = (начало_действия_лицензии_date + timedelta(days=365 * 3)).strftime('%Y-%m-%d')
            if оплачено:
                остаток = 0
            elif счёт_списания == '12' and not оплачено:
                остаток = итоговая_стоимость
            elif счёт_списания == '36' and not оплачено:
                окончание_действия_лицензии_date = datetime.strptime(окончание_действия_лицензии, '%Y-%m-%d').date()
                срок_действия_мес = (окончание_действия_лицензии_date - date.today()).days / 30
                мес_стоимость = итоговая_стоимость / 36
                остаток = round(мес_стоимость * срок_действия_мес, 2)
            else:
                остаток = 0
            cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute(""" UPDATE лицензии
                        SET
                        наименование_ПО=%s,
                        вендор=%s,
                        начало_действия_лицензии=%s,
                        окончание_действия_лицензии=%s,
                        счёт_списания=%s,
                        стоимость_за_единицу=%s,
                        итоговая_стоимость=%s,
                        заказчик_ПО=%s,
                        признак_ПО=%s,
                        количество_ПО=%s,
                        срок_действия_лицензии=GREATEST(to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW(), interval '0 seconds'),
                        оплачено=%s,
                        остаток=%s,
                        примечание=%s
                        WHERE номер_пп=%s
                        """, (наименование_ПО, вендор, начало_действия_лицензии, окончание_действия_лицензии, счёт_списания, стоимость_за_единицу, итоговая_стоимость, заказчик_ПО, признак_ПО, количество_ПО, окончание_действия_лицензии, оплачено, остаток, примечание, id))
            flash("Запись успешно обновлена!")
            conn2.commit()
            return redirect(url_for('home_editor'))
        else:
            flash('Такого ПО нет в справочнике ПО!')
            return redirect(url_for('editor_get_licence', id=id))
          
@app.route('/delete', methods=['POST', 'GET'])
def delete_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('delete_checkbox'):
            cur.execute('DELETE FROM лицензии WHERE код = %s', (id, ))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('home'))

@app.route('/editor_delete', methods=['POST', 'GET'])
def editor_delete_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_delete_checkbox'):
            cur.execute('DELETE FROM лицензии WHERE код = %s', (id, ))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('home_editor'))
    
@app.route('/delete_all_info', methods=['POST', 'GET'])
def delete_all_info():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM лицензии')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('home'))

@app.route('/editor_delete_all_info', methods=['POST', 'GET'])
def editor_delete_all_info():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM лицензии')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('home_editor'))
                
@app.route('/profile')
def profile(): 
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
   
    if 'loggedin' in session:
        cursor.execute('SELECT * FROM users WHERE id = %s', [session['id']])
        account = cursor.fetchone()
        return render_template('profile.html', account=account)
    return redirect(url_for('login'))

@app.route('/support_profile')
def support_profile():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        cur.execute('SELECT * FROM users WHERE id=%s', [session['id']])
        account = cur.fetchone()
        return render_template('support_profile.html', account=account)
    return redirect(url_for('login'))

@app.route('/editor_profile')
def editor_profile():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        cur.execute('SELECT * FROM users WHERE id=%s', [session['id']])
        account = cur.fetchone()
        return render_template('editor_profile.html', account=account)
    return redirect(url_for('login'))

@app.route('/types')
@role_required('admin')
def types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = 'SELECT * FROM Виды_ПО'
        cur.execute(string)
        list_types = cur.fetchall()
        return render_template('types.html', list_types=list_types)
    return redirect(url_for('login'))

@app.route('/editor_types')
@role_required('editor')
def editor_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = 'SELECT * FROM Виды_ПО'
        cur.execute(string)
        list_types = cur.fetchall()
        return render_template('editor_types.html', list_types = list_types)
    return redirect(url_for('login'))

@app.route('/support_types')
@role_required('support')
def support_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = 'SELECT * FROM Виды_ПО'
        cur.execute(string)
        list_types = cur.fetchall()
        return render_template('support_types.html', list_type=list_types)
    return redirect(url_for('login'))

@app.route('/show_types', methods=['POST', 'GET'])
def show_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, вид_ПО FROM Виды_ПО')
    list_types = cur.fetchall()
    return render_template('show_types.html', list_types=list_types)

@app.route('/editor_show_types', methods=['POST', 'GET'])
def editor_show_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, вид_ПО FROM Виды_ПО')
    list_types = cur.fetchall()
    return render_template('editor_show_types.html', list_types=list_types)

@app.route('/add_types', methods=['POST'])
def add_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        вид_ПО = request.form['вид_ПО']
        cur.execute("INSERT INTO Виды_ПО (наименование_ПО, вид_ПО) VALUES(%s,%s)", (наименование_ПО, вид_ПО))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('types'))
    
@app.route('/editor_add_types', methods=['POST'])
def editor_add_types():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        вид_ПО = request.form['вид_ПО']
        cur.execute('INSERT INTO Виды_ПО (наименование_ПО, вид_ПО) VALUES (%s, %s)', (наименование_ПО, вид_ПО))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_types'))
    
 
@app.route('/software')
@role_required('admin')
def software_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_ПО'
        cur.execute(string)
        list_software = cur.fetchall()
        return render_template('software.html', list_software=list_software)
    return redirect(url_for('login'))

@app.route('/show_software', methods=['GET', 'POST'])
def show_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_ПО')
    soft = cur.fetchall()
    return render_template('show_software.html', soft=soft)

@app.route('/editor_software')
@role_required('editor')
def editor_software_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_ПО'
        cur.execute(string)
        list_software = cur.fetchall()
        return render_template('editor_software.html', list_software=list_software)
    return redirect(url_for('login'))

@app.route('/editor_show_software', methods=['GET', 'POST'])
def editor_show_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_ПО')
    soft = cur.fetchall()
    return render_template('editor_add_software.html', soft=soft)

@app.route('/support_software')
@role_required('support')
def support_software_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = 'SELECT * FROM Справочник_ПО'
        cur.execute(string)
        list_software = cur.fetchall()
        return render_template('support_software.html', list_software=list_software)
    return redirect(url_for('login'))

@app.route('/add_software', methods=['POST'])
def add_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_ПО = request.form['код_ПО']
        наименование_ПО = request.form['наименование_ПО']
        описание_ПО = request.form['описание_ПО']
        ссылка_на_сайт_ПО = request.form['ссылка_на_сайт_ПО']
        вендор = request.form['вендор']
        стоимость_за_единицу = request.form['стоимость_за_единицу']
        признак_ПО = request.form.get('признак_ПО')
        примечание = request.form['примечание']
        cur.execute('SELECT наименование_ПО, признак_ПО FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s', (наименование_ПО, признак_ПО))
        soft = cur.fetchone()
        if soft:
            flash('Запись с такими данными уже существует!')
            return render_template('add_software.html')
        cur.execute("INSERT INTO Справочник_ПО (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)", (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('software_list'))

@app.route('/editor_add_software', methods=['POST'])
def editor_add_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_ПО = request.form['код_ПО']
        наименование_ПО = request.form['наименование_ПО']
        описание_ПО = request.form['описание_ПО']
        ссылка_на_сайт_ПО = request.form['ссылка_на_сайт_ПО']
        вендор = request.form['вендор']
        стоимость_за_единицу = request.form['стоимость_за_единицу']
        признак_ПО = request.form.get('признак_ПО')
        примечание = request.form['примечание']
        cur.execute('SELECT наименование_ПО, признак_ПО FROM Справочник_ПО WHERE наименование_ПО=%s AND признак_ПО=%s;', (наименование_ПО, признак_ПО,))
        soft = cur.fetchone()
        if soft:
            flash('Запись с такими данными уже существует!')
            return render_template('editor_add_software.html')
        cur.execute("INSERT INTO Справочник_ПО (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)", (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_software_list'))

@app.route('/edit_software/<id>', methods=['POST', 'GET'])
def edit_software(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_ПО WHERE код_ПО=%s', (id,))
    get_software = cur.fetchall()
    cur.close()
    return render_template('edit_software.html', software=get_software[0])

@app.route('/editor_edit_software/<id>', methods=['POST', 'GET'])
def editor_edit_software(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_ПО WHERE код_ПО=%s', (id,))
    get_software = cur.fetchall()
    cur.close()
    return render_template('editor_edit_software.html', software=get_software[0])

@app.route('/update_software/<id>', methods=['POST'])
def update_software(id):
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        описание_ПО = request.form['описание_ПО']
        ссылка_на_сайт_ПО = request.form['ссылка_на_сайт_ПО']
        вендор = request.form['вендор']
        стоимость_за_единицу = request.form['стоимость_за_единицу']
        признак_ПО = request.form['признак_ПО']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_ПО
                    SET
                    наименование_ПО=%s,
                    описание_ПО=%s,
                    ссылка_на_сайт_ПО=%s,
                    вендор=%s,
                    стоимость_за_единицу=%s,
                    признак_ПО=%s,
                    примечание=%s
                    WHERE код_ПО=%s
                    """, (наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('software_list'))
    
@app.route('/editor_update_software/<id>', methods=['POST'])
def editor_update_software(id):
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        описание_ПО = request.form['описание_ПО']
        ссылка_на_сайт_ПО = request.form['ссылка_на_сайт_ПО']
        вендор = request.form['вендор']
        стоимость_за_единицу = request.form['стоимость_за_единицу']
        признак_ПО = request.form['признак_ПО']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_ПО
                    SET
                    наименование_ПО=%s,
                    описание_ПО=%s,
                    ссылка_на_сайт_ПО=%s,
                    вендор=%s,
                    стоимость_за_единицу=%s,
                    признак_ПО=%s,
                    примечание=%s
                    WHERE код_ПО=%s
                    """, (наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('editor_software_list'))
    
    
@app.route('/delete_software', methods=['POST', 'GET'])
def delete_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        for id in request.form.getlist('delete_software_checkbox'):
            cur.execute('DELETE FROM Справочник_ПО WHERE код_ПО = %s', (id, ))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('software_list'))
    
@app.route('/delete_all_software', methods=['POST', 'GET'])
def delete_all_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('software_list'))
    
@app.route('/editor_delete_all_software', methods=['POST', 'GET'])
def editor_delete_all_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('editor_software_list'))

@app.route('/editor_delete_software' ,methods=['POST', 'GET'])
def editor_delete_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_delete_software_checkbox'):
            cur.execute('DELETE FROM Справочник_ПО WHERE код_ПО = %s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_software_list'))

@app.route('/download_software/report/excel')
def download_software_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_ПО')
    res_software = cur.fetchall()
    output_software = io.BytesIO()
    workbook_software = xlwt.Workbook()
    sh_software = workbook_software.add_sheet('Отчет по ПО')
    sh_software.write(0, 0, 'Код ПО')
    sh_software.write(0, 1, 'Наименование ПО')
    sh_software.write(0, 2, 'Описание ПО')
    sh_software.write(0, 3, 'Ссылка на сайт ПО')
    sh_software.write(0, 4, 'Вендор')
    sh_software.write(0, 5, 'Стоимость за единицу')
    sh_software.write(0, 6, 'Признак ПО')
    sh_software.write(0, 7, 'Примечание')
    idx = 0 
    for row in res_software:
        sh_software.write(idx+1, 0, str(row['код_ПО']))
        sh_software.write(idx+1, 1, row['наименование_ПО'])
        sh_software.write(idx+1, 2, row['описание_ПО'])
        sh_software.write(idx+1, 3, row['ссылка_на_сайт_ПО'])
        sh_software.write(idx+1, 4, row['вендор'])
        sh_software.write(idx+1, 5, row['стоимость_за_единицу'])
        sh_software.write(idx+1, 6, row['признак_ПО'])
        sh_software.write(idx+1, 7, row['примечание'])
        idx += 1
    workbook_software.save(output_software)
    output_software.seek(0)
    return Response(output_software, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=software_report.xls"})

@app.route('/vendor')
@role_required('admin')
def vendor_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_производителей_ПО'
        cur.execute(string)
        list_vendor = cur.fetchall()
        return render_template('vendor.html', list_vendor=list_vendor)
    return redirect(url_for('login'))

@app.route('/editor_vendor')
@role_required('editor')
def editor_vendor_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_производителей_ПО'
        cur.execute(string)
        list_vendor = cur.fetchall()
        return render_template('editor_vendor.html', list_vendor=list_vendor)
    return redirect(url_for('login'))

@app.route('/support_vendor')
@role_required('support')
def support_vendor_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_производителей_ПО'
        cur.execute(string)
        list_vendor = cur.fetchall()
        return render_template('support_vendor.html', list_vendor=list_vendor)
    return redirect(url_for('login'))

@app.route('/show_vendor', methods=['GET', 'POST'])
def show_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_производителей_ПО')
    list_vendor = cur.fetchall()
    return render_template('add_vendor.html', list_vendor=list_vendor)

@app.route('/editor_show_vendor', methods=['GET', 'POST'])
def editor_show_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_производителей_ПО')
    list_vendor = cur.fetchall()
    return render_template('editor_add_vendor.html', list_vendor=list_vendor)

@app.route('/add_vendor', methods=['POST'])
def add_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_производителя = request.form['код_производителя']
        производитель = request.form['производитель']
        описание_производителя = request.form['описание_производителя']
        ссылка_на_сайт_производителя = request.form['ссылка_на_сайт_производителя']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Справочник_производителей_ПО (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание) VALUES(%s,%s,%s,%s,%s)', (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('vendor_list'))
    
@app.route('/editor_add_vendor', methods=['POST'])
def editor_add_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_производителя = request.form['код_производителя']
        производитель = request.form['производитель']
        описание_производителя = request.form['описание_производителя']
        ссылка_на_сайт_производителя = request.form['ссылка_на_сайт_производителя']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Справочник_производителей_ПО (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание) VALUES(%s,%s,%s,%s,%s)', (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_vendor_list'))

@app.route('/edit_vendor/<id>', methods=['POST', 'GET'])
def edit_vendor(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_производителей_ПО WHERE код_производителя=%s', (id,))
    vendors = cur.fetchall()
    cur.close()
    return render_template('edit_vendor.html', vendor = vendors[0])

@app.route('/editor_edit_vendor/<id>', methods=['POST', 'GET'])
def editor_edit_vendor(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_производителей_ПО WHERE код_производителя=%s', (id,))
    vendors = cur.fetchall()
    cur.close()
    return render_template('editor_edit_vendor.html', vendor = vendors[0])

@app.route('/update_vendor/<id>', methods=['POST'])
def update_vendor(id):
    if request.method == 'POST':
        производитель = request.form['производитель']
        описание_производителя = request.form['описание_производителя']
        ссылка_на_сайт_производителя = request.form['ссылка_на_сайт_производителя']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_производителей_ПО
                    SET
                    производитель=%s,
                    описание_производителя=%s,
                    ссылка_на_сайт_производителя=%s,
                    примечание=%s
                    WHERE код_производителя=%s
                    """, (производитель, описание_производителя, ссылка_на_сайт_производителя, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('vendor_list'))
    
@app.route('/editor_update_vendor/<id>', methods=['POST'])
def editor_update_vendor(id):
    if request.method == 'POST':
        производитель = request.form['производитель']
        описание_производителя = request.form['описание_производителя']
        ссылка_на_сайт_производителя = request.form['ссылка_на_сайт_производителя']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_производителей_ПО
                    SET
                    производитель=%s,
                    описание_производителя=%s,
                    ссылка_на_сайт_производителя=%s,
                    примечание=%s
                    WHERE код_производителя=%s
                    """, (производитель, описание_производителя, ссылка_на_сайт_производителя, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('editor_vendor_list'))

@app.route('/delete_vendor', methods=['POST', 'GET'])
def delete_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('add_vendor_checkbox'):
            cur.execute('DELETE FROM Справочник_производителей_ПО WHERE код_производителя=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('vendor_list'))

@app.route('/editor_delete_vendor', methods=['POST', 'GET'])
def editor_delete_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_add_vendor_checkbox'):
            cur.execute('DELETE FROM Справочник_производителей_ПО WHERE код_производителя=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_vendor_list'))
    
@app.route('/delete_all_vendor', methods=['POST', 'GET'])
def delete_all_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_производителей_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('vendor_list'))
    
@app.route('/editor_delete_all_vnedor', methods=['POST', 'GET'])
def editor_delete_all_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_производителей_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('editor_vendor_list'))
    
@app.route('/download_vendor/report/excel')
def download_vendor_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_производителей_ПО')
    result_vendor = cur.fetchall()
    output_vendor = io.BytesIO()
    workbook_vendor = xlwt.Workbook()
    sh_vendor = workbook_vendor.add_sheet('Отчет по производителям ПО')
    sh_vendor.write(0, 0, 'Код производителя')
    sh_vendor.write(0, 1, 'Производитель')
    sh_vendor.write(0, 2, 'Описание производителя')
    sh_vendor.write(0, 3, 'Ссылка на сайт производителя')
    sh_vendor.write(0, 4, 'Примечание')
    idx = 0
    for row in result_vendor:
        sh_vendor.write(idx+1, 0, str(row['код_производителя']))
        sh_vendor.write(idx+1, 1, row['производитель'])
        sh_vendor.write(idx+1, 2, row['описание_производителя'])
        sh_vendor.write(idx+1, 3, row['ссылка_на_сайт_производителя'])
        sh_vendor.write(idx+1, 4, row['примечание'])
        idx += 1
    workbook_vendor.save(output_vendor)
    output_vendor.seek(0)
    return Response(output_vendor, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=vendor_report.xls"})

@app.route('/customer')
@role_required('admin')
def customer_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание FROM Справочник_заказчиков_ПО'
        cur.execute(string)
        list_customer = cur.fetchall()
        return render_template('customer.html', list_customer=list_customer)
    return redirect(url_for('login'))

@app.route('/editor_customer')
@role_required('editor')
def editor_customer_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string= 'SELECT заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание FROM Справочник_заказчиков_ПО'
        cur.execute(string)
        list_customer = cur.fetchall()
        return render_template('editor_customer.html', list_customer=list_customer)
    return redirect(url_for('login'))

@app.route('/support_customer')
@role_required('support')
def support_customer_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание FROM Справочник_заказчиков_ПО'
        cur.execute(string)
        list_customer = cur.fetchall()
        return render_template('support_customer.html', list_customer=list_customer)
    return redirect(url_for('login'))

@app.route('/show_customer', methods=['GET', 'POST'])
def show_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание FROM Справочник_заказчиков_ПО')
    list_customer = cur.fetchall()
    return render_template('add_customer.html', list_customer=list_customer)

@app.route('/show_software_byid/<software>', methods=['POST', 'GET'])
def show_software_byid(software):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, вид_ПО FROM Виды_ПО WHERE наименование_ПО=%s', (software,))
    list_software = cur.fetchall()
    return render_template('software_byid.html', list_software=list_software)

@app.route('/editor_show_software_byid/<software>', methods=['POST', 'GET'])
def editor_show_software_byid(software):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_ПО, вид_ПО FROM Виды_ПО WHERE наименование_ПО=%s', (software,))
    list_software = cur.fetchall()
    return render_template('editor_software_byid.html', list_software=list_software)

@app.route('/show_customer_byid/<customer>', methods=['POST', 'GET'])
def show_customer_byid(customer):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты WHERE наименование_контрагента=%s', (customer,))
    list_customer = cur.fetchall()
    return render_template('customer_description.html', list_customer=list_customer, customer=customer)

@app.route('/editor_show_customer_byid/<customer>', methods=['POST', 'GET'])
def editor_show_customer_by_id(customer):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты WHERE наименование_контрагента=%s', (customer,))
    list_customer = cur.fetchall()
    return render_template('editor_customer_description.html', list_customer=list_customer, customer=customer)

@app.route('/editor_show_customer', methods=['GET', 'POST'])
def editor_show_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Контрагенты')
    list_customer=cur.fetchall()
    return render_template('editor_add_customer.html', list_customer=list_customer)

@app.route('/add_customer', methods=['POST'])
def add_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        заказчик_ПО = request.form['заказчик_ПО']
        описание_заказчика = request.form['описание_заказчика']
        ссылка_на_сайт_заказчика = request.form['ссылка_на_сайт_заказчика']
        примечание = request.form['примечание']
        cur.execute('SELECT COUNT(*) AS count FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (заказчик_ПО, ))
        cust = cur.fetchone()['count']
        if cust > 0:
            flash('Такой заказчик ПО уже существует!')
            return redirect(url_for('show_customer'))
        else:
            cur.execute('INSERT INTO Справочник_заказчиков_ПО (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание) VALUES(%s,%s,%s,%s)', (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('customer_list'))
    
@app.route('/editor_add_customer', methods=['POST'])
def editor_add_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        заказчик_ПО = request.form['заказчик_ПО']
        описание_заказчика = request.form['описание_заказчика']
        ссылка_на_сайт_заказчика = request.form['ссылка_на_сайт_заказчика']
        примечание = request.form['примечание']
        cur.execute('SELECT COUNT(*) AS count FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (заказчик_ПО, ))
        cust = cur.fetchone()['count']
        if cust > 0:
            flash('Такой заказчик ПО уже существует!')
            return redirect(url_for('editor_show_customer'))
        else:
            cur.execute('INSERT INTO Справочник_заказчиков_ПО (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание) VALUES(%s,%s,%s,%s)', (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_customer_list'))
    

@app.route('/edit_customer/<id>', methods=['POST','GET'])
def edit_customer(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (id,))
    customers = cur.fetchall()
    cur.close()
    return render_template('edit_customer.html', customer=customers[0])

@app.route('/editor_edit_customer/<id>', methods=['POST', 'GET'])
def editor_edit_customer(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (id,))
    customers = cur.fetchall()
    cur.close()
    return render_template('editor_edit_customer.html', customer=customers[0])

@app.route('/update_customer/<id>', methods=['POST'])
def update_customer(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        заказчик_ПО = request.form['заказчик_ПО']
        описание_заказчика = request.form['описание_заказчика']
        ссылка_на_сайт_заказчика = request.form['ссылка_на_сайт_заказчика']
        примечание = request.form['примечание']
        cur.execute(""" UPDATE Справочник_заказчиков_ПО
                    SET
                    заказчик_ПО=%s,
                    описание_заказчика=%s,
                    ссылка_на_сайт_заказчика=%s,
                    примечание=%s
                    WHERE код_заказчика=%s
                    """, (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('customer_list'))
    
@app.route('/editor_update_customer/<id>', methods=['POST'])
def editor_update_customer(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        заказчик_ПО = request.form['заказчик_ПО']
        описание_заказчика = request.form['описание_заказчика']
        ссылка_на_сайт_заказчика = request.form['ссылка_на_сайт_заказчика']
        примечание = request.form['примечание']
        cur.execute(""" UPDATE Справочник_заказчиков_ПО
                    SET
                    заказчик_ПО=%s,
                    описание_заказчика=%s,
                    ссылка_на_сайт_заказчика=%s,
                    примечание=%s
                    WHERE код_заказчика=%s
                    """, (заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание, id))
        flash("Запись успешно обновлена!")
        conn2.commit()
        return redirect(url_for('editor_customer_list'))
    
@app.route('/delete_customer', methods=['POST', 'GET'])
def delete_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('add_customer_checkbox'):
            cur.execute('DELETE FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('customer_list'))
    
@app.route('/delete_all_customer', methods=['POST', 'GET'])
def delete_all_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_заказчиков_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('customer_list'))
    
@app.route('/editor_delete_all_customer', methods=['POST', 'GET'])
def editor_delete_all_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_заказчиков_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('editor_customer_list'))

@app.route('/editor_delete_customer', methods=['POST', 'GET'])
def editor_delete_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_add_customer_checkbox'):
            cur.execute('DELETE FROM Справочник_заказчиков_ПО WHERE заказчик_ПО=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_customer_list'))

@app.route('/download_customer/report/excel')
def download_customer_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_заказчиков_ПО')
    result_customer = cur.fetchall()
    output_customer = io.BytesIO()
    workbook_customer = xlwt.Workbook()
    sh_customer = workbook_customer.add_sheet('Отчет по заказчикам')
    sh_customer.write(0, 0, 'Заказчик ПО')
    sh_customer.write(0, 1, 'Описание заказчика')
    sh_customer.write(0, 2, 'Ссылка на сайт заказчика')
    sh_customer.write(0, 3, 'Примечание')
    idx = 0
    for row in result_customer:
        sh_customer.write(idx+1, 0, row['заказчик_ПО'])
        sh_customer.write(idx+1, 1, row['описание_заказчика'])
        sh_customer.write(idx+1, 2, row['ссылка_на_сайт_заказчика'])
        sh_customer.write(idx+1, 3, row['примечание'])
        idx += 1
    workbook_customer.save(output_customer)
    output_customer.seek(0)
    return Response(output_customer, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=customer_report.xls"})

@app.route('/licence')
@role_required('admin')
def licence_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_лицензий'
        cur.execute(string)
        list_licence = cur.fetchall()
        return render_template('licence.html', list_licence=list_licence)
    return redirect(url_for('login'))

@app.route('/editor_licence')
@role_required('editor')
def editor_licence_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_лицензий'
        cur.execute(string)
        list_licence = cur.fetchall()
        return render_template('editor_licence.html', list_licence=list_licence)
    return redirect(url_for('login'))

@app.route('/support_licence')
@role_required('support')
def support_licence_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string='SELECT * FROM Справочник_лицензий'
        cur.execute(string)
        list_licence = cur.fetchall()
        return render_template('support_licence.html', list_licence=list_licence)
    return redirect(url_for('login'))

@app.route('/show_licence_list', methods=['GET', 'POST'])
def show_licence_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_лицензий')
    list_licence = cur.fetchall()
    return render_template('add_licence_to_list.html', list_licence=list_licence)

@app.route('/editor_show_licence_list', methods=['GET', 'POST'])
def editor_show_licence_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_лицензий')
    list_licence = cur.fetchall()
    return render_template('editor_add_licence_to_list.html', list_licence=list_licence)

    
@app.route('/add_licence_to_list', methods=['POST'])
def add_licence_to_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_лицензии = request.form['код_лицензии']
        наименование_лицензии = request.form['наименование_лицензии']
        тип_лицензии = request.form['тип_лицензии']
        счёт_списания = request.form['счёт_списания']
        версия_лицензии = request.form['версия_лицензии']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Справочник_лицензий (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание) VALUES(%s,%s,%s,%s,%s,%s)', (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('licence_list'))

@app.route('/editor_add_licence_to_list', methods=['POST'])
def editor_add_licence_to_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        код_лицензии = request.form['код_лицензии']
        наименование_лицензии = request.form['наименование_лицензии']
        тип_лицензии = request.form['тип_лицензии']
        счёт_списания = request.form['счёт_списания']
        версия_лицензии = request.form['версия_лицензии']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Справочник_лицензий (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание) VALUES(%s,%s,%s,%s,%s,%s)', (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_licence_list'))
    
@app.route('/edit_licence/<id>', methods=['POST', 'GET'])
def edit_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_лицензий WHERE код_лицензии=%s', (id))
    lic = cur.fetchall()
    return render_template('edit_licence.html', lice = lic[0])

@app.route('/editor_edit_licence/<id>', methods=['POST', 'GET'])
def editor_edit_licence(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_лицензий WHERE код_лицензии=%s', (id))
    lic = cur.fetchall()
    return render_template('editor_edit_licence.html', lice = lic[0])

@app.route('/update_licence/<id>', methods=['POST'])
def update_licence_list(id):
    if request.method == 'POST':
        наименование_лицензии = request.form['наименование_лицензии']
        тип_лицензии = request.form['тип_лицензии']
        счёт_списания = request.form.get('счёт_списания')
        версия_лицензии = request.form['версия_лицензии']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_лицензий
                    SET наименование_лицензии=%s,
                    тип_лицензии=%s,
                    счёт_списания=%s,
                    версия_лицензии=%s,
                    примечание=%s
                    WHERE код_лицензии=%s
                    """, (наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание, id))
        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('licence_list'))
    
@app.route('/editor_update_licence/<id>', methods=['POST'])
def editor_update_licence_list(id):
    if request.method == 'POST':
        наименование_лицензии = request.form['наименование_лицензии']
        тип_лицензии = request.form['тип_лицензии']
        счёт_списания = request.form.get('счёт_списания')
        версия_лицензии = request.form['версия_лицензии']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Справочник_лицензий
                    SET наименование_лицензии=%s,
                    тип_лицензии=%s,
                    счёт_списания=%s,
                    версия_лицензии=%s,
                    примечание=%s
                    WHERE код_лицензии=%s
                    """, (наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание, id))
        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('editor_licence_list'))
    
@app.route('/delete_licence', methods=['POST', 'GET'])
def delete_licence_from_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('add_licence_to_list_checkbox'):
            cur.execute('DELETE FROM Справочник_лицензий WHERE код_лицензии=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('licence_list'))

@app.route('/editor_delete_licence', methods=['POST', 'GET'])
def editor_delete_licence_from_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_add_licence_to_list_checkbox'):
            cur.execute('DELETE FROM Справочник_лицензий WHERE код_лицензии=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_licence_list'))
    
@app.route('/delete_all_licence', methods=['POST', 'GET'])
def delete_all_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_лицензий')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('licence_list'))
    
@app.route('/editor_delete_all_licence', methods=['POST', 'GET'])
def editor_delete_all_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Справочник_лицензий')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('editor_licence_list'))

@app.route('/partners_list')
@role_required('admin')
def partners_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = "SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты"
        cur.execute(string)
        partner_list = cur.fetchall()
        return render_template('partner.html', partner_list=partner_list)
    return redirect(url_for('login'))

@app.route('/editor_partners_list')
@role_required('editor')
def editor_partners_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = "SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты"
        cur.execute(string)
        partner_list = cur.fetchall()
        return render_template('editor_partner.html', partner_list=partner_list)
    return redirect(url_for('login'))

@app.route('/support_partners_list')
@role_required('support')
def support_partners_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string="SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты"
        cur.execute(string)
        partner_list = cur.fetchall()
        return render_template('support_partner.html', partner_list=partner_list)
    return redirect(url_for('login'))

@app.route('/show_partners_list', methods=['GET', 'POST'])
def show_partners_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты")
    partner_list = cur.fetchall()
    return render_template('show_partners_list.html', partner_list=partner_list)

@app.route('/editor_show_partners_list', methods=['GET', 'POST'])
def editor_show_partners_list():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты")
    partner_list = cur.fetchall()
    return render_template('editor_show_partners_list.html', partner_list=partner_list)

@app.route('/add_partner', methods=['POST'])
def add_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_контрагента = request.form['наименование_контрагента']
        краткое_наименование_предприятия = request.form['краткое_наименование_предприятия']
        полное_наименование_предприятия = request.form['полное_наименование_предприятия']
        ИНН = request.form['ИНН']
        КПП = request.form['КПП']
        юридический_адрес = request.form['юридический_адрес']
        фактический_адрес = request.form['фактический_адрес']
        ОКПО = request.form['ОКПО']
        ОГРН = request.form['ОГРН']
        телефон = request.form['телефон']
        email = request.form['email']
        ссылка_на_сайт_контрагента = request.form['ссылка_на_сайт_контрагента']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Контрагенты (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('partners_list'))
    
@app.route('/editor_add_partner', methods=['POST'])
def editor_add_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_контрагента = request.form['наименование_контрагента']
        краткое_наименование_предприятия = request.form['краткое_наименование_предприятия']
        полное_наименование_предприятия = request.form['полное_наименование_предприятия']
        ИНН = request.form['ИНН']
        КПП = request.form['КПП']
        юридический_адрес = request.form['юридический_адрес']
        фактический_адрес = request.form['фактический_адрес']
        ОКПО = request.form['ОКПО']
        ОГРН = request.form['ОГРН']
        телефон = request.form['телефон']
        email = request.form['email']
        ссылка_на_сайт_контрагента = request.form['ссылка_на_сайт_контрагента']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Контрагенты (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('editor_partners_list'))
    
@app.route('/edit_partner/<id>', methods=['POST', 'GET'])
def edit_partner(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты WHERE наименование_контрагента = %s', (id,))
    partners = cur.fetchall()
    return render_template('edit_partner.html', partner = partners[0])

@app.route('/editor_edit_partner/<id>', methods=['POST', 'GET'])
def editor_edit_partner(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание FROM Контрагенты WHERE наименование_контрагента =%s', (id,))
    partners=cur.fetchall()
    return render_template('editor_edit_partner.html', partner=partners[0])

@app.route('/update_partner/<id>', methods=['POST'])
def update_partner(id):
    if request.method == 'POST':
        наименование_контрагента = request.form['наименование_контрагента']
        краткое_наименование_предприятия = request.form['краткое_наименование_предприятия']
        полное_наименование_предприятия = request.form['полное_наименование_предприятия']
        ИНН = request.form['ИНН']
        КПП = request.form['КПП']
        юридический_адрес = request.form['юридический_адрес']
        фактический_адрес = request.form['фактический_адрес']
        ОКПО = request.form['ОКПО']
        ОГРН = request.form['ОГРН']
        телефон = request.form['телефон']
        email = request.form['email']
        ссылка_на_сайт_контрагента = request.form['ссылка_на_сайт_контрагента']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Контрагенты
                    SET наименование_контрагента=%s,
                    краткое_наименование_предприятия=%s,
                    полное_наименование_предприятия=%s,
                    ИНН=%s,
                    КПП=%s,
                    юридический_адрес=%s,
                    фактический_адрес=%s,
                    ОКПО=%s,
                    ОГРН=%s,
                    телефон=%s,
                    email=%s,
                    ссылка_на_сайт_контрагента=%s,
                    примечание=%s
                    WHERE код_контрагента=%s
                    """, (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание, id))
        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('partners_list'))
    
@app.route('/editor_update_partner/<id>', methods=['POST'])
def editor_update_partner(id):
    if request.method == 'POST':
        наименование_контрагента = request.form['наименование_контрагента']
        краткое_наименование_предприятия = request.form['краткое_наименование_предприятия']
        полное_наименование_предприятия = request.form['полное_наименование_предприятия']
        ИНН = request.form['ИНН']
        КПП = request.form['КПП']
        юридический_адрес = request.form['юридический_адрес']
        фактический_адрес = request.form['фактический_адрес']
        ОКПО = request.form['ОКПО']
        ОГРН = request.form['ОГРН']
        телефон = request.form['телефон']
        email = request.form['email']
        ссылка_на_сайт_контрагента = request.form['ссылка_на_сайт_контрагента']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Контрагенты
                    SET наименование_контрагента=%s,
                    краткое_наименование_предприятия=%s,
                    полное_наименование_предприятия=%s,
                    ИНН=%s,
                    КПП=%s,
                    юридический_адрес=%s,
                    фактический_адрес=%s,
                    ОКПО=%s,
                    ОГРН=%s,
                    телефон=%s,
                    email=%s,
                    ссылка_на_сайт_контрагента=%s,
                    примечание=%s
                    WHERE код_контрагента=%s
                    """, (наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН, телефон, email, ссылка_на_сайт_контрагента, примечание, id))
        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('editor_partners_list'))
    
@app.route('/delete_partner', methods=['POST', 'GET'])
def delete_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('delete_partners_checkbox'):
            cur.execute('DELETE FROM Контрагенты WHERE наименование_контрагента=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('partners_list'))

@app.route('/editor_delete_partner', methods=['POST', 'GET'])
def editor_delete_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_delete_partners_checkbox'):
            cur.execute('DELETE FROM Контрагенты WHERE наименование_контрагента=%s', (id,))
            conn2.commit()
            flash('Запись успешно удалена!')
            return redirect(url_for('editor_partners_list'))
        
@app.route('/delete_all_partner', methods=['POST', 'GET'])
def delete_all_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Контрагенты')
        conn2.commit()
    flash('Все записи были успешно удалены!')
    return redirect(url_for('partners_list'))

@app.route('/editor_delete_all_partner', methods=['POST', 'GET'])
def editor_delete_all_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Контрагенты')
        conn2.commit()
    flash('Все записи успешно удалены!')
    return redirect(url_for('editor_partners_list'))

@app.route('/download_partner/report/excel')
def download_partner_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Контрагенты')
    result_partner = cur.fetchall()
    output_partner = io.BytesIO()
    workbook_partner = xlwt.Workbook()
    sh_partner = workbook_partner.add_sheet('Отчет по контрагентам')
    sh_partner.write(0, 0, 'Код контрагента')
    sh_partner.write(0, 1, 'Наименование контрагента')
    sh_partner.write(0, 2, 'Краткое наименование предприятия')
    sh_partner.write(0, 3, 'Полное наименование предприятия')
    sh_partner.write(0, 4, 'ИНН')
    sh_partner.write(0, 5, 'КПП')
    sh_partner.write(0, 6, 'Юридический адрес')
    sh_partner.write(0, 7, 'Фактический адрес')
    sh_partner.write(0, 8, 'ОКПО')
    sh_partner.write(0, 9, 'ОГРН')
    sh_partner.write(0, 10, 'телефон')
    sh_partner.write(0, 11, 'email')
    sh_partner.write(0, 12, 'ссылка_на_сайт_контрагента')
    sh_partner.write(0, 13, 'Примечание')
    idx = 0
    for row in result_partner:
        sh_partner.write(idx+1, 0, str(row['код_контрагента']))
        sh_partner.write(idx+1, 1, row['наименование_контрагента'])
        sh_partner.write(idx+1, 2, row['краткое_наименование_организации'])
        sh_partner.write(idx+1, 3, row['полное_наименование_организации'])
        sh_partner.write(idx+1, 4, row['ИНН'])
        sh_partner.write(idx+1, 5, row['КПП'])
        sh_partner.write(idx+1, 6, row['юридический_адрес'])
        sh_partner.write(idx+1, 7, row['фактический_адрес'])
        sh_partner.write(idx+1, 8, row['ОКПО'])
        sh_partner.write(idx+1, 9, row['ОГРН'])
        sh_partner.write(idx+1, 10, row['телефон'])
        sh_partner.write(idx+1, 11, row['email'])
        sh_partner.write(idx+1, 12, row['ссылка_на_сайт_контрагента'])
        sh_partner.write(idx+1, 13, row['примечание'])
        idx += 1
    workbook_partner.save(output_partner)
    output_partner.seek(0)
    return Response(output_partner, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=partner_report.xls"})

@app.route('/download_licence/report/excel')
def download_licence_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Справочник_лицензий')
    result_licence = cur.fetchall()
    output_licence = io.BytesIO()
    workbook_licence = xlwt.Workbook()
    sh_licence = workbook_licence.add_sheet('Отчет по лицензиям')
    sh_licence.write(0, 0, 'Код лицензии')
    sh_licence.write(0, 1, 'Наименование лицензии')
    sh_licence.write(0, 2, 'Тип лицензии')
    sh_licence.write(0, 3, 'Счёт списания')
    sh_licence.write(0, 4, 'Версия лицензии')
    sh_licence.write(0, 5, 'Примечание')
    idx = 0  
    for row in result_licence:
        sh_licence.write(idx+1, 0, str(row['код_лицензии']))
        sh_licence.write(idx+1, 1, row['наименование_лицензии'])
        sh_licence.write(idx+1, 2, row['тип_лицензии'])
        sh_licence.write(idx+1, 3, row['счёт_списания'])
        sh_licence.write(idx+1, 4, row['версия_лицензии'])
        sh_licence.write(idx+1, 5, row['примечание'])
        idx += 1
    workbook_licence.save(output_licence)
    output_licence.seek(0)
    return Response(output_licence, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=licence_report.xls"})

@app.route('/download/report/excel')
def download_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM лицензии')
    result = cur.fetchall()
    output = io.BytesIO()
    workbook = xlwt.Workbook()
    sh = workbook.add_sheet('Отчет')
    sh.write(0, 0, 'Наименование ПО')
    sh.write(0, 1, 'Вендор')
    sh.write(0, 2, 'Начало действия лицензии')
    sh.write(0, 3, 'Окончание действия лицензии')
    sh.write(0, 4, 'Счёт списания')
    sh.write(0, 5, 'Стоимость за единицу')
    sh.write(0, 6, 'Итоговая стоимость')
    sh.write(0, 7, 'Заказчик ПО')
    sh.write(0, 8, 'Признак ПО')
    sh.write(0, 9, 'Количество ПО')
    sh.write(0, 10, 'Срок действия лицензии')
    sh.write(0, 11, 'Оплачено')
    sh.write(0, 12, 'Остаток')
    sh.write(0, 13, 'Примечание')
    sh.write(0, 14, 'Номер пп')
    idx = 0
    for row in result:
        sh.write(idx+1, 0, row['наименование_ПО'])
        sh.write(idx+1, 1, row['вендор'])
        sh.write(idx+1, 2, row['начало_действия_лицензии'])
        sh.write(idx+1, 3, row['окончание_действия_лицензии'])
        sh.write(idx+1, 4, row['счёт_списания'])
        sh.write(idx+1, 5, row['стоимость_за_единицу'])
        sh.write(idx+1, 6, row['итоговая_стоимость'])
        sh.write(idx+1, 7, row['заказчик_ПО'])
        sh.write(idx+1, 8, row['признак_ПО'])
        sh.write(idx+1, 9, row['количество_ПО'])
        sh.write(idx+1, 10, str(row['срок_действия_лицензии']))
        sh.write(idx+1, 11, row['оплачено'])
        sh.write(idx+1, 12, row['остаток'])
        sh.write(idx+1, 13, row['примечание'])
        sh.write(idx+1, 14, str(row['номер_пп']))
        idx += 1
    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=total_report.xls"})

@app.route('/contracts')
@role_required('admin')
def contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT * FROM Договоры'
    cur.execute(string)
    contract_list = cur.fetchall()
    return render_template('contract.html', contract_list=contract_list)

@app.route('/editor_contracts')
@role_required('editor')
def editor_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT * FROM Договоры'
    cur.execute(string)
    contract_list = cur.fetchall()
    return render_template('editor_contract.html', contract_list=contract_list)

@app.route('/support_contracts')
@role_required('support')
def support_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT * FROM Договоры'
    cur.execute(string)
    contract_list = cur.fetchall()
    return render_template('support_contract.html', contract_list=contract_list)

@app.route('/show_contracts', methods=['GET', 'POST'])
def show_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание FROM Договоры'
    cur.execute(string)
    contract_list = cur.fetchall()
    return render_template('add_contracts.html', contract_list=contract_list)

@app.route('/editor_show_contract', methods=['GET', 'POST'])
def editor_show_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание FROM Договоры'
    cur.execute(string)
    contract_list = cur.fetchall()
    return render_template('editor_add_contracts.html', contract_list=contract_list)

@app.route('/add_contracts', methods=['POST'])
def add_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        тип_договора = request.form['тип_договора']
        регистрационный_номер = request.form['регистрационный_номер']
        контрагент = request.form['контрагент']
        статус = request.form['статус']
        предмет_договора = request.form['предмет_договора']
        дата_договора = request.form['дата_договора']
        дата_начала_действия_договора = request.form['дата_начала_действия_договора']
        дата_окончания_действия_договора = request.form['дата_окончания_действия_договора']
        сумма_без_ндс = request.form['сумма_без_ндс']
        сумма_с_ндс = request.form['сумма_с_ндс']
        филиал = request.form['филиал']
        примечание = request.form['примечание']
        cur.execute("""INSERT INTO Договоры (тип_договора, регистрационный_номер, контрагент, статус, предмет_договора,
                    дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс,
                    сумма_с_ндс, филиал, примечание) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
        тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора,
        дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание,))
        conn2.commit()
        flash('Запись была успешно создана!')
        return redirect(url_for('contracts'))
    
@app.route('/editor_add_contract', methods=['POST'])
def editor_add_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        тип_договора = request.form['тип_договора']
        регистрационный_номер = request.form['регистрационный_номер']
        контрагент = request.form['контрагент']
        статус = request.form['статус']
        предмет_договора = request.form['предмет_договора']
        дата_договора = request.form['дата_договора']
        дата_начала_действия_договора = request.form['дата_начала_действия_договора']
        дата_окончания_действия_договора = request.form['дата_окончания_действия_договора']
        сумма_без_ндс = request.form['сумма_без_ндс']
        сумма_с_ндс = request.form['сумма_с_ндс']
        филиал = request.form['филиал']
        примечание = request.form['примечание']
        cur.execute("""INSERT INTO Договоры (тип_договора, регистрационный_номер, контрагент, статус, предмет_договора,
                    дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс,
                    сумма_с_ндс, филиал, примечание) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
        тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора,
        дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание,))
        conn2.commit()
        flash('Запись была успешно создана!')
        return redirect(url_for('editor_contracts'))
    
@app.route('/edit_contract/<id>', methods=['GET', 'POST'])
def edit_contract(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Договоры WHERE код_договора=%s', (id,))
    contract_list = cur.fetchall()
    return render_template('edit_contracts.html', contract_list = contract_list[0])

@app.route('/editor_edit_contract/<id>', methods=['GET', 'POST'])
def editor_edit_contract(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Договоры WHERE код_договора=%s', (id,))
    contract_list = cur.fetchall()
    return render_template('editor_edit_contracts.html', contract_list = contract_list[0])

@app.route('/update_contracts/<id>', methods=['POST'])
def update_contracts(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        тип_договора = request.form['тип_договора']
        регистрационный_номер = request.form['регистрационный_номер']
        контрагент = request.form['контрагент']
        статус = request.form['статус']
        предмет_договора = request.form['предмет_договора']
        дата_договора = request.form['дата_договора']
        дата_начала_действия_договора = request.form['дата_начала_действия_договора']
        дата_окончания_действия_договора = request.form['дата_окончания_действия_договора']
        сумма_без_ндс = request.form['сумма_без_ндс']
        сумма_с_ндс = request.form['сумма_с_ндс']
        филиал = request.form['филиал']
        примечание = request.form['примечание']
        cur.execute("""UPDATE Договоры
                    SET тип_договора=%s, регистрационный_номер=%s, контрагент=%s, статус=%s, предмет_договора=%s,
                    дата_договора=%s, дата_начала_действия_договора=%s, дата_окончания_действия_договора=%s, сумма_без_ндс=%s,
                    сумма_с_ндс=%s, филиал=%s, примечание=%s WHERE код_договора=%s""", (тип_договора, регистрационный_номер, контрагент, статус,
                    предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс,
                    сумма_с_ндс, филиал, примечание, id,))
        conn2.commit()
        flash('Запись была успешно изменена!')
        return redirect(url_for('contracts'))
    
@app.route('/editor_update_contracts/<id>', methods=['POST'])
def editor_update_contracts(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        тип_договора = request.form['тип_договора']
        регистрационный_номер = request.form['регистрационный_номер']
        контрагент = request.form['контрагент']
        статус = request.form['статус']
        предмет_договора = request.form['предмет_договора']
        дата_договора = request.form['дата_договора']
        дата_начала_действия_договора = request.form['дата_начала_действия_договора']
        дата_окончания_действия_договора = request.form['дата_окончания_действия_договора']
        сумма_без_ндс = request.form['сумма_без_ндс']
        сумма_с_ндс = request.form['сумма_с_ндс']
        филиал = request.form['филиал']
        примечание = request.form['примечание']
        cur.execute("""UPDATE Договоры
                    SET тип_договора=%s, регистрационный_номер=%s, контрагент=%s, статус=%s, предмет_договора=%s,
                    дата_договора=%s, дата_начала_действия_договора=%s, дата_окончания_действия_договора=%s, сумма_без_ндс=%s,
                    сумма_с_ндс=%s, филиал=%s, примечание=%s WHERE код_договора=%s""", (тип_договора, регистрационный_номер, контрагент, статус,
                    предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс,
                    сумма_с_ндс, филиал, примечание, id,))
        conn2.commit()
        flash('Запись была успешно изменена!')
        return redirect(url_for('editor_contracts'))
    
@app.route('/delete_contract',methods=['GET', 'POST'])
def delete_contract():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('delete_contracts_checkbox'):
            cur.execute('DELETE FROM Договоры WHERE код_договора=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('contracts'))
    
@app.route('/editor_delete_contract', methods=['GET', 'POST'])
def editor_delete_contract():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('editor_delete_contracts_checkbox'):
            cur.execute('DELETE FROM Договоры WHERE код_договора=%s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('editor_contracts'))
    
@app.route('/delete_all_contract', methods=['POST', 'GET'])
def delete_all_contract():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Договоры')
        conn2.commit()
    flash('Все записи успешно удалены!')
    return redirect(url_for('contracts'))

@app.route('/editor_delete_all_contracts', methods=['POST', 'GET'])
def editor_delete_all_contracts():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Договоры')
        conn2.commit()
    flash('Все записи были успешно удалены!')
    return redirect(url_for('editor_contracts'))

@app.route('/download_contract/report/excel')
def download_contract_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Договоры')
    result_contract = cur.fetchall()
    output_contract = io.BytesIO()
    workbook_contract = xlwt.Workbook()
    sh_contract = workbook_contract.add_sheet('Отчет по договорам')
    sh_contract.write(0, 0, 'Код договора')
    sh_contract.write(0, 1, 'Тип договора')
    sh_contract.write(0, 2, 'Регистрационный номер')
    sh_contract.write(0, 3, 'Контрагент')
    sh_contract.write(0, 4, 'Статус')
    sh_contract.write(0, 5, 'Предмет договора')
    sh_contract.write(0, 6, 'Дата договора')
    sh_contract.write(0, 7, 'Дата начала действия договора')
    sh_contract.write(0, 8, 'Дата окончания действия договора')
    sh_contract.write(0, 9, 'Сумма без НДС')
    sh_contract.write(0, 10, 'Сумма с НДС')
    sh_contract.write(0, 11, 'Филиал')
    sh_contract.write(0, 12, 'Примечание')
    idx = 0  
    for row in result_contract:
        sh_contract.write(idx+1, 0, str(row['код_договора']))
        sh_contract.write(idx+1, 1, row['тип_договора'])
        sh_contract.write(idx+1, 2, row['регистрационный_номер'])
        sh_contract.write(idx+1, 3, row['контрагент'])
        sh_contract.write(idx+1, 4, row['статус'])
        sh_contract.write(idx+1, 5, row['предмет_договора'])
        sh_contract.write(idx+1, 6, row['дата_договора'])
        sh_contract.write(idx+1, 7, row['дата_начала_действия_договора'])
        sh_contract.write(idx+1, 8, row['дата_окончания_действия_договора'])
        sh_contract.write(idx+1, 9, row['сумма_без_ндс'])
        sh_contract.write(idx+1, 10, row['сумма_с_ндс'])
        sh_contract.write(idx+1, 11, row['филиал'])
        sh_contract.write(idx+1, 12, row['примечание'])
        idx += 1
    workbook_contract.save(output_contract)
    output_contract.seek(0)
    return Response(output_contract, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=contract_report.xls"})

@app.route('/upload_contract', methods=['POST', 'GET'])
def upload_contract():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=5, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    cur.execute(
                        "INSERT INTO Договоры (код_договора, тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        (row[0], row[3], row[4], row[17], row[12], row[11], row[6], row[8], row[9], row[34], row[35], row[1], row[29])
                    )
                    cur.execute('SELECT COUNT(*) FROM Контрагенты WHERE наименование_контрагента=%s', (row[17],))
                    count = cur.fetchone()[0]
                    if count == 0:
                        cur.execute('INSERT INTO Контрагенты(наименование_контрагента, телефон, email, ИНН, ОКПО, ОГРН) VALUES (%s, %s, %s, %s, %s, %s)', (row[17], row[30], row[31], row[18], row[20], row[19]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('contracts'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('contracts'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('contracts'))
        return render_template('contract.html')
    else:
        return render_template('contract.html')  
    
@app.route('/editor_upload_contract', methods=['POST', 'GET'])
def editor_upload_contract():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    if len(row) == 15: 
                        data.append(row)
                    else:
                        flash('Неправильный формат файла')
                        return redirect(url_for('editor_contracts'))
                for row in data:
                    cur.execute(
                        "INSERT INTO Договоры (код_договора, тип_договора, регистрационный_номер, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, сумма_без_ндс, сумма_с_ндс, филиал, примечание) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12])
                    )
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_contracts'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_contracts'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_contracts'))
        return render_template('editor_contract.html')
    else:
        return render_template('editor_contract.html')  
    
@app.route('/support_install_software')
@role_required('support')
def support_install_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT код_установки, наименование_ПО, тип_лицензии, чекбокс, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание FROM Установка_ПО'
    cur.execute(string)
    list_installation = cur.fetchall()
    return render_template('support_install.html', list_installation=list_installation)

@app.route('/editor_installation_software')
def editor_installation_software():
    if 'role' in session and session['role'] == 'editor':
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        string='SELECT * FROM Установка_ПО'
        cur.execute(string)
        list_installation = cur.fetchall()
        return render_template('editor_install.html', list_installation=list_installation)
    return redirect(url_for('login'))

@app.route('/install_software')
@role_required('admin')
def install_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    string='SELECT код_установки, наименование_ПО, тип_лицензии, чекбокс, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание FROM Установка_ПО'
    cur.execute(string)
    list_installation = cur.fetchall()
    return render_template('install.html', list_installation=list_installation)

@app.route('/show_installation', methods=['GET', 'POST'])
def show_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Установка_ПО')
    list_installation = cur.fetchall()
    return render_template('add_install.html', list_installation=list_installation)

@app.route('/support_show_installation', methods=['GET', 'POST'])
def support_show_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Установка_ПО')
    list_installation = cur.fetchall()
    return render_template('support_add_install.html', list_installation=list_installation)

@app.route('/admin_add_installation', methods=['POST'])
def admin_add_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form.get('тип_лицензии')
        ФИО = request.form['ФИО']
        отдел = request.form['отдел']
        ip_адрес = request.form['ip_адрес']
        наименование_машины = request.form['наименование_машины']
        чекбокс = bool(request.form.get('чекбокс'))
        дата_установки_ПО = request.form.get('дата_установки_ПО')
        if not чекбокс:
            дата_установки_ПО = None
        if not дата_установки_ПО:
            дата_установки_ПО = None
        else:
            try:
                дата_установки_ПО = datetime.strptime(дата_установки_ПО, '%Y-%m-%d').date()
            except ValueError:
                flash('Некорректный формат даты!', 'warning')
                return redirect(url_for('install_software'))
        чекбокс_условно_бесплатное_ПО = bool(request.form.get('чекбокс_условно_бесплатное_ПО'))
        примечание = request.form['примечание']
        if чекбокс_условно_бесплатное_ПО:
            тип_лицензии = 'Условно-бесплатная'
        cur.execute("""SELECT количество_лицензий_ПО FROM Учет_лицензий
                    WHERE наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии))
        licence_data = cur.fetchone()
    
        if licence_data:
            общее_количество = licence_data['количество_лицензий_ПО']
        else:
            общее_количество = 0  
        cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
                    наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 

        число_установленных_лицензий = cur.fetchone()['count']

        if чекбокс:
            число_установленных_лицензий += 1
        
        if чекбокс and дата_установки_ПО is None:
            flash('Введите дату установки ПО!')
            return redirect(url_for('install_software'))
        if дата_установки_ПО and дата_установки_ПО > datetime.now().date():
           flash('Дата установки ПО не может быть больше текущей даты!', 'warning')
           return redirect(url_for('install_software'))
        if not чекбокс_условно_бесплатное_ПО and число_установленных_лицензий > общее_количество:
           flash('Превышено количество доступных лицензий!', 'warning')
           return redirect(url_for('install_software'))
        else:
            cur.execute('INSERT INTO Установка_ПО (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание))
            conn2.commit()
            flash('Запись успешно создана!')
            return redirect(url_for('install_software'))
        
@app.route('/add_installation', methods=['POST'])
def add_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form.get('тип_лицензии')
        ФИО = request.form['ФИО']
        отдел = request.form['отдел']
        ip_адрес = request.form['ip_адрес']
        наименование_машины = request.form['наименование_машины']
        чекбокс = bool(request.form.get('чекбокс'))
        дата_установки_ПО = request.form.get('дата_установки_ПО')
        if not чекбокс:
            дата_установки_ПО = None
        else:
            try:
                дата_установки_ПО = datetime.strptime(дата_установки_ПО, '%Y-%m-%d').date()
            except ValueError:
                flash('Введен некорректный формат даты', 'warning')
                return redirect(url_for('support_install_software'))
        чекбокс_условно_бесплатное_ПО = bool(request.form.get('чекбокс_условно_бесплатное_ПО'))
        примечание = request.form['примечание']
        cur.execute("""SELECT количество_лицензий_ПО FROM Учет_лицензий
                    WHERE наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии))
        licence_data = cur.fetchone()
    
        if licence_data:
            общее_количество = licence_data['количество_лицензий_ПО']
        else:
            общее_количество = 0  
        cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
                    наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 
        число_установленных_лицензий = cur.fetchone()['count']
        if чекбокс: 
            число_установленных_лицензий += 1
            
        if чекбокс and дата_установки_ПО is None:
            flash('Введите дату установки ПО!')
            return redirect(url_for('support_install_software'))
        if дата_установки_ПО and дата_установки_ПО > datetime.now().date():
           flash('Дата установки ПО не может быть больше текущей даты!', 'warning')
           return redirect(url_for('support_install_software'))
        if not чекбокс_условно_бесплатное_ПО and число_установленных_лицензий > общее_количество:
           flash('Превышено количество доступных лицензий!', 'warning')
           return redirect(url_for('support_install_software'))
        else:
            cur.execute('INSERT INTO Установка_ПО (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание))
            conn2.commit()
            flash('Запись успешно создана!')
            return redirect(url_for('support_install_software'))
    
@app.route('/edit_installation/<id>', methods=['POST', 'GET'])
def edit_installation(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Установка_ПО WHERE код_установки=%s', (id,))
    installations = cur.fetchall()
    return render_template('edit_installation.html', installation = installations[0])

@app.route('/admin_edit_installation/<id>', methods=['POST', 'GET'])
def admin_edit_installation(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM Установка_ПО WHERE код_установки=%s", (id,))
    installations = cur.fetchall()
    return render_template('admin_edit_installation.html', installation = installations[0])
                    
@app.route('/update_installation/<id>', methods=['POST'])
def update_installation(id):
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form['тип_лицензии']
        ФИО = request.form['ФИО']
        отдел = request.form['отдел']
        ip_адрес = request.form['ip_адрес']
        наименование_машины = request.form['наименование_машины']
        чекбокс = bool(request.form.get('чекбокс'))
        дата_установки_ПО = request.form['дата_установки_ПО']
        чекбокс_условно_бесплатное_ПО = bool(request.form.get('чекбокс_условно_бесплатное_ПО'))
        if чекбокс_условно_бесплатное_ПО:
            тип_лицензии = 'Условно-бесплатная'
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""SELECT наименование_ПО, тип_лицензии, число_установленных_лицензий , чекбокс
                    FROM Установка_ПО WHERE код_установки=%s""", (id,))
        installation_data = cur.fetchone() #данные из бд
        
        if not installation_data:
            flash('Установка с указанным ID не найдена!')
            return redirect(url_for('edit_installation', id=id))
        
        cur.execute("""SELECT количество_лицензий_ПО FROM Учет_лицензий
                    WHERE наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии))
        new_licence_data = cur.fetchone() #общее количество лицензий ПО

        if new_licence_data:
            общее_количество = new_licence_data['количество_лицензий_ПО']
        else:
            общее_количество = 0
        
        if installation_data['чекбокс'] != чекбокс:
            if чекбокс: #False to True
                if installation_data['тип_лицензии'] != тип_лицензии:
                    cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
                  наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 
                    число_установленных_лицензий = cur.fetchone()['count']
                    if число_установленных_лицензий + 1 > общее_количество:
                        flash('Превышено допустимое количество лицензий!')
                        return redirect(url_for('edit_installation', id=id))
                    if installation_data['число_установленных_лицензий'] >0:
                        cur.execute(""" UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий - 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s
                                    """, (наименование_ПО, installation_data['тип_лицензии']))    
                        cur.execute(""" UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий + 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s
                                    """, (наименование_ПО, тип_лицензии))
                        cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                        max_id = cur.fetchone()[0]

                        new_id = max_id + 1

                        cur.execute("""
                             INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                             VALUES (%s, %s, %s, 1)
                         """, (new_id, наименование_ПО, тип_лицензии))
                        flash('Запись успешно обновлена!')
                        conn2.commit()
                        return redirect(url_for('support_install_software'))
                    else:
                        cur.execute("""UPDATE Установка_ПО
                                    SET число_установленных_лицензий = 0
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, installation_data['тип_лицензии']))
                        cur.execute("""UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий + 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                        cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                        max_id = cur.fetchone()[0]

                        new_id = max_id + 1

                        cur.execute("""
                             INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                             VALUES (%s, %s, %s, 1)
                         """, (new_id, наименование_ПО, тип_лицензии))
                        flash('Запись успешно обновлена!')
                        conn2.commit()
                        return redirect(url_for('support_install_software'))
                    
                cur.execute(""" UPDATE Установка_ПО
                            SET наименование_ПО=%s,
                            тип_лицензии=%s,
                            ФИО=%s,
                            отдел=%s,
                            ip_адрес=%s,
                            наименование_машины=%s,
                            чекбокс=%s,
                            общее_количество=%s,
                            дата_установки_ПО=%s,
                            чекбокс_условно_бесплатное_ПО=%s,
                            примечание=%s
                            WHERE код_установки=%s
                            """, (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины,
                                   чекбокс, общее_количество,
                                   дата_установки_ПО, 
                                   чекбокс_условно_бесплатное_ПО, примечание, id))

                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('support_install_software'))
            else: #True to False
                if installation_data['число_установленных_лицензий'] > 0:
                    cur.execute("""UPDATE Установка_ПО SET число_установленных_лицензий = число_установленных_лицензий - 1
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                    flash('Запись успешно обновлена!')
                    conn2.commit()
                    return redirect(url_for('support_install_software'))
                else:
                    cur.execute("""UPDATE Установка_ПО SET число_установленных_лицензий = 0
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                    flash('Запись успешно обновлена!')
                    conn2.commit()
                    return redirect(url_for('support_install_software'))

        if not чекбокс:
            дата_установки_ПО = None
            cur.execute("""UPDATE Установка_ПО
                        SET число_установленных_лицензий = число_установленных_лицензий WHERE наименование_ПО=%s AND 
                        тип_лицензии=%s;""", (наименование_ПО, installation_data['тип_лицензии']))
            flash('Запись успешно обновлена!')
            conn2.commit()
            return redirect(url_for('support_install_software'))
        else:
            try:
                дата_установки_ПО = datetime.strptime(дата_установки_ПО, '%Y-%m-%d').date()
            except ValueError:
                flash('Введен некорректный формат даты', 'warning')
                return redirect(url_for('edit_installation', id=id))
            
        if чекбокс and дата_установки_ПО is None:
            flash('Введите дату установки ПО!')
            return redirect(url_for('edit_installation', id=id))
        if дата_установки_ПО and дата_установки_ПО > datetime.now().date():
           flash('Дата установки ПО не может быть больше текущей даты!', 'warning')
           return redirect(url_for('edit_installation', id=id))
        
       
        if installation_data['тип_лицензии'] != тип_лицензии:
            cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
          наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 
            число_установленных_лицензий = cur.fetchone()['count']
            if число_установленных_лицензий + 1 > общее_количество:
                flash('Превышено допустимое количество лицензий!')
                return redirect(url_for('edit_installation', id=id))
            if installation_data['число_установленных_лицензий'] > 0:
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий - 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, installation_data['тип_лицензии']))    
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий + 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, тип_лицензии))
                cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                max_id = cur.fetchone()[0]

                new_id = max_id + 1

                cur.execute("""
                     INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                     VALUES (%s, %s, %s, 1)
                 """, (new_id, наименование_ПО, тип_лицензии))
                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('support_install_software'))
            else:
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = 0
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, installation_data['тип_лицензии']))    
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий + 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, тип_лицензии))
                cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                max_id = cur.fetchone()[0]

                new_id = max_id + 1

                cur.execute("""
                     INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                     VALUES (%s, %s, %s, 1)
                 """, (new_id, наименование_ПО, тип_лицензии))
                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('support_install_software'))
            
        cur.execute(""" UPDATE Установка_ПО
                    SET наименование_ПО=%s,
                    тип_лицензии=%s,
                    ФИО=%s,
                    отдел=%s,
                    ip_адрес=%s,
                    наименование_машины=%s,
                    чекбокс=%s,
                    общее_количество=%s,
                    дата_установки_ПО=%s,
                    чекбокс_условно_бесплатное_ПО=%s,
                    примечание=%s
                    WHERE код_установки=%s
                    """, (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины,
                           чекбокс, общее_количество,
                           дата_установки_ПО, 
                           чекбокс_условно_бесплатное_ПО, примечание, id))

        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('support_install_software'))
    
@app.route('/admin_update_installation/<id>', methods=['POST'])
def admin_update_installation(id):
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form['тип_лицензии']
        ФИО = request.form['ФИО']
        отдел = request.form['отдел']
        ip_адрес = request.form['ip_адрес']
        наименование_машины = request.form['наименование_машины']
        чекбокс = bool(request.form.get('чекбокс'))
        дата_установки_ПО = request.form['дата_установки_ПО']
        чекбокс_условно_бесплатное_ПО = bool(request.form.get('чекбокс_условно_бесплатное_ПО'))
        if чекбокс_условно_бесплатное_ПО:
            тип_лицензии = 'Условно-бесплатная'
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""SELECT наименование_ПО, тип_лицензии, число_установленных_лицензий , чекбокс
                    FROM Установка_ПО WHERE код_установки=%s""", (id,))
        installation_data = cur.fetchone() #данные из бд
        
        if not installation_data:
            flash('Установка с указанным ID не найдена!')
            return redirect(url_for('admin_edit_installation', id=id))
        
        cur.execute("""SELECT количество_лицензий_ПО FROM Учет_лицензий
                    WHERE наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии))
        new_licence_data = cur.fetchone() #общее количество лицензий ПО

        if new_licence_data:
            общее_количество = new_licence_data['количество_лицензий_ПО']
        else:
            общее_количество = 0
        
        if installation_data['чекбокс'] != чекбокс:
            if чекбокс: #False to True
                if installation_data['тип_лицензии'] != тип_лицензии:
                    cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
                  наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 
                    число_установленных_лицензий = cur.fetchone()['count']
                    if число_установленных_лицензий + 1 > общее_количество:
                        flash('Превышено допустимое количество лицензий!')
                        return redirect(url_for('admin_edit_installation', id=id))
                    if installation_data['число_установленных_лицензий'] > 0:
                        cur.execute(""" UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий - 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s
                                    """, (наименование_ПО, installation_data['тип_лицензии']))    
                        cur.execute(""" UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий + 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s
                                    """, (наименование_ПО, тип_лицензии))
                        cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                        max_id = cur.fetchone()[0]

                        new_id = max_id + 1

                        cur.execute("""
                             INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                             VALUES (%s, %s, %s, 1)
                         """, (new_id, наименование_ПО, тип_лицензии))
                        flash('Запись успешно обновлена!')
                        conn2.commit()
                        return redirect(url_for('install_software'))
                    else:
                        cur.execute("""UPDATE Установка_ПО
                                    SET число_установленных_лицензий = 0
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, installation_data['тип_лицензии']))
                        cur.execute("""UPDATE Установка_ПО
                                    SET число_установленных_лицензий = число_установленных_лицензий + 1
                                    WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                        cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                        max_id = cur.fetchone()[0]

                        new_id = max_id + 1

                        cur.execute("""
                             INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                             VALUES (%s, %s, %s, 1)
                         """, (new_id, наименование_ПО, тип_лицензии))
                        flash('Запись успешно обновлена!')
                        conn2.commit()
                        return redirect(url_for('install_software'))
                    
                cur.execute(""" UPDATE Установка_ПО
                            SET наименование_ПО=%s,
                            тип_лицензии=%s,
                            ФИО=%s,
                            отдел=%s,
                            ip_адрес=%s,
                            наименование_машины=%s,
                            чекбокс=%s,
                            общее_количество=%s,
                            дата_установки_ПО=%s,
                            чекбокс_условно_бесплатное_ПО=%s,
                            примечание=%s
                            WHERE код_установки=%s
                            """, (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины,
                                   чекбокс, общее_количество,
                                   дата_установки_ПО, 
                                   чекбокс_условно_бесплатное_ПО, примечание, id))

                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('install_software'))
            else: #True to False
                if installation_data['число_установленных_лицензий'] > 0:
                    cur.execute("""UPDATE Установка_ПО SET число_установленных_лицензий = число_установленных_лицензий - 1
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                    flash('Запись успешно обновлена!')
                    conn2.commit()
                    return redirect(url_for('install_software'))
                else:
                    cur.execute("""UPDATE Установка_ПО SET число_установленных_лицензий = 0
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                    flash('Запись успешно обновлена!')
                    conn2.commit()
                    return redirect(url_for('install_software'))

        if not чекбокс:
            дата_установки_ПО = None
            cur.execute("""UPDATE Установка_ПО
                        SET число_установленных_лицензий = число_установленных_лицензий WHERE наименование_ПО=%s AND 
                        тип_лицензии=%s;""", (наименование_ПО, installation_data['тип_лицензии']))
            flash('Запись успешно обновлена!')
            conn2.commit()
            return redirect(url_for('install_software'))
        else:
            try:
                дата_установки_ПО = datetime.strptime(дата_установки_ПО, '%Y-%m-%d').date()
            except ValueError:
                flash('Введен некорректный формат даты', 'warning')
                return redirect(url_for('admin_edit_installation', id=id))
            
        if чекбокс and дата_установки_ПО is None:
            flash('Введите дату установки ПО!')
            return redirect(url_for('edit_installation', (id,)))
        if дата_установки_ПО and дата_установки_ПО > datetime.now().date():
           flash('Дата установки ПО не может быть больше текущей даты!', 'warning')
           return redirect(url_for('admin_edit_installation', id=id))
        
       
        if installation_data['тип_лицензии'] != тип_лицензии:
            cur.execute("""SELECT COUNT(*) AS count FROM Установка_ПО WHERE 
          наименование_ПО=%s AND тип_лицензии=%s""", (наименование_ПО, тип_лицензии)) 
            число_установленных_лицензий = cur.fetchone()['count']
            if число_установленных_лицензий + 1 > общее_количество:
                flash('Превышено допустимое количество лицензий!')
                return redirect(url_for('admin_edit_installation', id=id))
            if installation_data['число_установленных_лицензий'] > 0:
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий - 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, installation_data['тип_лицензии']))    
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий + 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, тип_лицензии))
                cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                max_id = cur.fetchone()[0]

                new_id = max_id + 1

                cur.execute("""
                    INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                    VALUES (%s, %s, %s, 1)
                """, (new_id, наименование_ПО, тип_лицензии))
                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('install_software'))
            else:
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = 0
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, installation_data['тип_лицензии']))    
                cur.execute(""" UPDATE Установка_ПО
                            SET число_установленных_лицензий = число_установленных_лицензий + 1
                            WHERE наименование_ПО=%s AND тип_лицензии=%s
                            """, (наименование_ПО, тип_лицензии))
                cur.execute("SELECT MAX(код_установки) FROM Установка_ПО")
                max_id = cur.fetchone()[0]

                new_id = max_id + 1

                cur.execute("""
                     INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, число_установленных_лицензий)
                     VALUES (%s, %s, %s, 1)
                 """, (new_id, наименование_ПО, тип_лицензии))
                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('install_software'))
        
                flash('Запись успешно обновлена!')
                conn2.commit()
                return redirect(url_for('install_software'))
            
        cur.execute(""" UPDATE Установка_ПО
                    SET наименование_ПО=%s,
                    тип_лицензии=%s,
                    ФИО=%s,
                    отдел=%s,
                    ip_адрес=%s,
                    наименование_машины=%s,
                    чекбокс=%s,
                    общее_количество=%s,
                    дата_установки_ПО=%s,
                    чекбокс_условно_бесплатное_ПО=%s,
                    примечание=%s
                    WHERE код_установки=%s
                    """, (наименование_ПО, тип_лицензии, ФИО, отдел, ip_адрес, наименование_машины,
                           чекбокс, общее_количество,
                           дата_установки_ПО, 
                           чекбокс_условно_бесплатное_ПО, примечание, id))

        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('install_software'))
    
@app.route('/delete_installation', methods=['POST', 'GET'])
def delete_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('support_delete_installation_checkbox'):
            cur.execute('SELECT наименование_ПО, тип_лицензии, число_установленных_лицензий FROM Установка_ПО WHERE код_установки=%s', (id,))
            installation_info = cur.fetchone()
            if installation_info:
                наименование_ПО = installation_info['наименование_ПО']
                тип_лицензии = installation_info['тип_лицензии']
                число_установленных_лицензий = installation_info['число_установленных_лицензий']
                if число_установленных_лицензий > 0:
                    cur.execute(
                        """
                        UPDATE Установка_ПО
                        SET число_установленных_лицензий = число_установленных_лицензий - 1 
                        WHERE наименование_ПО=%s AND тип_лицензии=%s;
                        """, (наименование_ПО, тип_лицензии,))
                    cur.execute('DELETE FROM Установка_ПО WHERE код_установки=%s',(id,))
                    conn2.commit()
            
                    flash('Запись успешно удалена!')
                else:
                    cur.execute("""UPDATE Установка_ПО
                                SET число_установленных_лицензий = 0
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии))
                    cur.execute('DELETE FROM Установка_ПО WHERE код_установки=%s', (id,))
                    conn2.commit()
                    flash('Запись успешно удалена!')
            else:
                flash('Запись не найдена!')
    
        return redirect(url_for('support_install_software'))

@app.route('/admin_delete_installation', methods=['POST', 'GET'])
def admin_delete_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('admin_delete_installation_checkbox'):
            cur.execute('SELECT наименование_ПО, тип_лицензии, число_установленных_лицензий FROM Установка_ПО WHERE код_установки=%s', (id,))
            installation_info = cur.fetchone()
            if installation_info:
                число_установленных_лицензий = installation_info['число_установленных_лицензий']
                наименование_ПО = installation_info['наименование_ПО']
                тип_лицензии = installation_info['тип_лицензии']
                if число_установленных_лицензий > 0:
                    cur.execute('DELETE FROM Установка_ПО WHERE код_установки=%s', (id,))
                    cur.execute(
                        """
                        UPDATE Установка_ПО
                        SET число_установленных_лицензий = число_установленных_лицензий - 1 
                        WHERE наименование_ПО=%s AND тип_лицензии=%s;
                        """,
                        (наименование_ПО, тип_лицензии,),
                    )
                    conn2.commit()
            
                    flash('Запись успешно удалена!')
                else:
                    cur.execute('DELETE FROM Установка_ПО WHERE код_установки=%s', (id,))
                    cur.execute("""UPDATE Установка_ПО
                                SET число_установленных_лицензий = 0
                                WHERE наименование_ПО=%s AND тип_лицензии=%s;""", (наименование_ПО, тип_лицензии,))
                    conn2.commit()
                    flash('Запись успешно удалена!')
            else:
                flash('Запись не найдена!')
        return redirect(url_for('install_software'))
    
@app.route('/support_delete_all_install', methods=['POST', 'GET'])
def support_delete_all_install():
    cur = conn2.cursor(curso_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Установка_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('support_install_software'))
    
@app.route('/delete_all_install', methods=['POST', 'GET'])
def delete_all_install():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        cur.execute('DELETE FROM Установка_ПО')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('install_software'))
    
@app.route('/download/software_install_report/excel')
def download_software_install_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Установка_ПО')
    result = cur.fetchall()
    output = io.BytesIO()
    workbook = xlwt.Workbook()
    sh = workbook.add_sheet('Отчет')
    sh.write(0, 0, 'Код установки')
    sh.write(0, 1, 'Наименование ПО')
    sh.write(0, 2, 'Тип лицензии')
    sh.write(0, 3, 'ФИО')
    sh.write(0, 4, 'Отдел')
    sh.write(0, 5, 'IP адрес')
    sh.write(0, 6, 'Наименование машины')
    sh.write(0, 7, 'Чекбокс')
    sh.write(0, 8, 'Общее количество')
    sh.write(0, 9, 'Число установленных лицензий')
    sh.write(0, 10, 'Дата установки ПО')
    sh.write(0, 11, 'Чекбокс условно-бесплатное ПО')
    sh.write(0, 12, 'Примечание')
    idx = 0
    for row in result:
        sh.write(idx+1, 0, str(row['код_установки']))
        sh.write(idx+1, 1, row['наименование_ПО'])
        sh.write(idx+1, 2, row['тип_лицензии'])
        sh.write(idx+1, 3, row['ФИО'])
        sh.write(idx+1, 4, row['отдел'])
        sh.write(idx+1, 5, row['ip_адрес'])
        sh.write(idx+1, 6, row['наименование_машины'])
        sh.write(idx+1, 7, row['чекбокс'])
        sh.write(idx+1, 8, row['общее_количество'])
        sh.write(idx+1, 9, row['число_установленных_лицензий'])
        sh.write(idx+1, 10, row['дата_установки_ПО'])
        sh.write(idx+1, 11, row['чекбокс_условно_бесплатное_ПО'])
        sh.write(idx+1, 12, row['примечание'])
        idx += 1
    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=software_installation_report.xls"})

@app.route('/number_licences')
@role_required('admin')
def number_licences():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = ('SELECT * FROM Учет_лицензий')
        cur.execute(string)
        list_number = cur.fetchall()
        return render_template('number.html', list_number=list_number)
    return redirect(url_for('login'))

@app.route('/editor_number_licences')
@role_required('editor')
def editor_number_licences():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = 'SELECT * FROM Учет_лицензий'
        cur.execute(string)
        list_number = cur.fetchall()
        return render_template('editor_number.html', list_number=list_number)
    return redirect(url_for('login'))

@app.route('/view_number_licences')
@role_required('support')
def view_number_licences():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'loggedin' in session:
        string = ('SELECT * FROM Учет_лицензий')
        cur.execute(string)
        list_number = cur.fetchall()
        return render_template('support_number.html', list_number=list_number)
    return redirect(url_for('login'))

@app.route('/show_number_licences', methods=['GET', 'POST'])
def show_number_licences():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Учет_лицензий')
    list_number = cur.fetchall()
    return render_template('add_number_licences.html', list_number=list_number)

@app.route('/add_number', methods=['POST'])
def add_number():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        номер_заявки = request.form['номер_заявки']
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form.get('тип_лицензии')
        количество_лицензий_ПО = request.form['количество_лицензий_ПО']
        примечание = request.form['примечание']
        cur.execute('INSERT INTO Учет_лицензий (номер_заявки, наименование_ПО, тип_лицензии, количество_лицензий_ПО, примечание) VALUES(%s,%s,%s,%s,%s)', (номер_заявки, наименование_ПО, тип_лицензии, количество_лицензий_ПО, примечание))
        conn2.commit()
        flash('Запись успешно создана!')
        return redirect(url_for('number_licences'))
    

@app.route('/edit_number/<id>', methods=['POST', 'GET'])
def edit_number(id):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Учет_лицензий WHERE номер_заявки=%s', (id,))
    numbers = cur.fetchall()
    return render_template('edit_number.html', number = numbers[0])

@app.route('/update_number/<id>', methods=['POST'])
def update_number(id):
    if request.method == 'POST':
        наименование_ПО = request.form['наименование_ПО']
        тип_лицензии = request.form.get('тип_лицензии')
        количество_лицензий_ПО = request.form['количество_лицензий_ПО']
        примечание = request.form['примечание']
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(""" UPDATE Учет_лицензий
                    SET наименование_ПО=%s,
                    тип_лицензии=%s,
                    количество_лицензий_ПО=%s,
                    примечание=%s
                    WHERE номер_заявки=%s
                    """, (наименование_ПО, тип_лицензии, количество_лицензий_ПО, примечание, id))
        flash('Запись успешно обновлена!')
        conn2.commit()
        return redirect(url_for('number_licences'))
    
@app.route('/delete_number', methods=['POST', 'GET'])
def delete_number():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        for id in request.form.getlist('delete_number_checkbox'):
            cur.execute('DELETE FROM Учет_лицензий WHERE номер_заявки= %s', (id,))
            conn2.commit()
        flash('Запись успешно удалена!')
        return redirect(url_for('number_licences'))
    
@app.route('/delete_all_number', methods=['POST', 'GET'])
def delete_all_number():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        cur.execute('DELETE FROM Учет_лицензий')
        conn2.commit()
        flash('Все записи были успешно удалены!')
        return redirect(url_for('number_licences'))

@app.route('/download/number_report/excel')
def download_number_report():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT * FROM Учет_лицензий')
    result = cur.fetchall()
    output = io.BytesIO()
    workbook = xlwt.Workbook()
    sh = workbook.add_sheet('Отчет')
    sh.write(0, 0, 'Номер заявки')
    sh.write(0, 1, 'Наименование ПО')
    sh.write(0, 2, 'Тип лицензии')
    sh.write(0, 3, 'Количество лицензий ПО')
    sh.write(0, 4, 'Примечание')
    idx = 0
    for row in result:
        sh.write(idx+1, 0, str(row['номер_заявки']))
        sh.write(idx+1, 1, row['наименование_ПО'])
        sh.write(idx+1, 2, row['тип_лицензии'])
        sh.write(idx+1, 3, row['количество_лицензий_ПО'])
        sh.write(idx+1, 4, row['примечание'])
        idx += 1
    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=number_report.xls"})

filtered_contracts = []

@app.route('/submit_contract', methods=['POST', 'GET'])
def submit_contract():
    global filtered_contracts  
    if request.method == 'POST':
        код_договора = request.form.get('код_договора')
        контрагент = request.form.get('контрагент')
        статус = request.form.get('статус')
        предмет_договора = request.form.get('предмет_договора')
        дата_договора = request.form.get('дата_договора')
        дата_начала_действия_договора = request.form.get('дата_начала_действия_договора')
        дата_окончания_действия_договора = request.form.get('дата_окончания_действия_договора')
        филиал = request.form.get('филиал')

        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT код_договора, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, филиал FROM Договоры WHERE "
        params = []

        if код_договора:
            query += "код_договора ILIKE %s AND "
            params.append(код_договора)
        if контрагент:
            query += "контрагент ILIKE %s AND "
            params.append(контрагент)
        if статус:
            query += "статус ILIKE %s AND "
            params.append(статус)
        if предмет_договора:
            query += "предмет_договора ILIKE %s AND "
            params.append(f'%{предмет_договора}%')
        if дата_договора:
            query += "дата_договора = %s AND "
            params.append(дата_договора)
        if дата_начала_действия_договора:
            query += "дата_начала_действия_договора = %s AND "
            params.append(дата_начала_действия_договора)
        if дата_окончания_действия_договора:
            query += "дата_окончания_действия_договора = %s AND "
            params.append(дата_окончания_действия_договора)
        if филиал:
            query += "филиал ILIKE %s AND "
            params.append(f'%{филиал}%')

        if len(params) > 0:
            query = query[:-5] 
        else:
            return render_template('update_contract.html', filtered_contracts=[])

        cur.execute(query, params)
        filtered_contracts = cur.fetchall()  
        cur.close()

        return render_template('update_contract.html', filtered_contracts=filtered_contracts)

    return render_template('update_contract.html', filtered_contracts=[])

editor_filtered_contracts = []
@app.route('/editor_submit_contract', methods=['POST', 'GET'])
def editor_submit_contract():
    global editor_filtered_contracts  
    if request.method == 'POST':
        код_договора = request.form.get('код_договора')
        контрагент = request.form.get('контрагент')
        статус = request.form.get('статус')
        предмет_договора = request.form.get('предмет_договора')
        дата_договора = request.form.get('дата_договора')
        дата_начала_действия_договора = request.form.get('дата_начала_действия_договора')
        дата_окончания_действия_договора = request.form.get('дата_окончания_действия_договора')
        филиал = request.form.get('филиал')

        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT код_договора, контрагент, статус, предмет_договора, дата_договора, дата_начала_действия_договора, дата_окончания_действия_договора, филиал FROM Договоры WHERE "
        editor_params = []

        if код_договора:
            query += "код_договора ILIKE %s AND "
            editor_params.append(код_договора)
        if контрагент:
            query += "контрагент ILIKE %s AND "
            editor_params.append(контрагент)
        if статус:
            query += "статус ILIKE %s AND "
            editor_params.append(статус)
        if предмет_договора:
            query += "предмет_договора ILIKE %s AND "
            editor_params.append(f'%{предмет_договора}%')
        if дата_договора:
            query += "дата_договора = %s AND "
            editor_params.append(дата_договора)
        if дата_начала_действия_договора:
            query += "дата_начала_действия_договора = %s AND "
            editor_params.append(дата_начала_действия_договора)
        if дата_окончания_действия_договора:
            query += "дата_окончания_действия_договора = %s AND "
            editor_params.append(дата_окончания_действия_договора)
        if филиал:
            query += "филиал ILIKE %s AND "
            editor_params.append(f'%{филиал}%')

        if len(editor_params) > 0:
            query = query[:-5] 
        else:
            return render_template('editor_update_contract.html', filtered_contracts=[])

        cur.execute(query, editor_params)
        editor_filtered_contracts = cur.fetchall()  
        cur.close()

        return render_template('editor_update_contract.html', editor_filtered_contracts=editor_filtered_contracts)

    return render_template('editor_update_contract.html', editor_filtered_contracts=[])

@app.route('/export_contracts', methods=['POST', 'GET'])
def export_contracts():
    global filtered_contracts 
    if filtered_contracts:
        df = pd.DataFrame(filtered_contracts)
        excel_file_path = 'contracts.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404 
    
@app.route('/editor_export_contracts', methods=['POST', 'GET'])
def editor_export_contracts():
    global editor_filtered_contracts
    if editor_filtered_contracts:
        df = pd.DataFrame(editor_filtered_contracts)
        excel_file_path = 'contracts.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404
    
filtered_total = []
@app.route('/submit_total', methods=['POST', 'GET'])
def submit_total():
    global filtered_total  
    if request.method == 'POST':
        статья_затрат = request.form.get('статья_затрат')
        наименование_ПО_БУ = request.form.get('наименование_ПО_БУ')
        наименование_ПО = request.form.get('наименование_ПО')
        краткое_наименование_ПО = request.form.get('краткое_наименование_ПО')
        код = request.form.get('код')
        филиал = request.form.get('филиал')
        счёт_затрат = request.form.get('счёт_затрат')
        вид_деятельности = request.form.get('вид_деятельности')
        срок_полезного_использования_мес = request.form.get('срок_полезного_использования_мес')
        дата_начала_списания = request.form.get('дата_начала_списания')
        дата_окончания_списания = request.form.get('дата_окончания_списания')
        договор_счет = request.form.get('договор_счет')
        контрагент = request.form.get('контрагент')
        первичный_документ = request.form.get('первичный_документ')
        страна_производитель = request.form.get('страна_производитель')
        правообладатель = request.form.get('правообладатель')
        срок_предоставления_права = request.form.get('срок_предоставления_права')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, код, филиал, счет_затрат, вид_деятельности, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, срок_предоставления_права FROM общая_информация WHERE "
        params_total = []
        if статья_затрат:
            query += "статья_затрат = %s AND "
            params_total.append(f'%{статья_затрат}%')
        if наименование_ПО_БУ:
            query += "наименование_ПО_БУ = %s AND "
            params_total.append(f'%{наименование_ПО_БУ}%')
        if наименование_ПО:
            query += "наименование_ПО = %s AND "
            params_total.append(f'%{наименование_ПО}%')
        if краткое_наименование_ПО: 
            query += "краткое_наименование_ПО = %s AND "
            params_total.append(f'%{краткое_наименование_ПО}%')
        if код:
            query += "код = %s AND "
            params_total.append(f'%{код}%')
        if филиал:
            query += "филиал ILIKE %s AND "
            params_total.append(f'%{филиал}%')
        if счёт_затрат:
            query += "счёт_затрат = %s AND "
            params_total.append(счёт_затрат)
        if вид_деятельности:
            query += "вид_деятельности ILIKE %s AND "
            params_total.append(f'%{вид_деятельности}%')
        if срок_полезного_использования_мес:
            query += "срок_полезного_использования_мес = %s AND "
            params_total.append(срок_полезного_использования_мес)
        if дата_начала_списания:
            query += "дата_начала_списания = %s AND "
            params_total.append(дата_начала_списания)
        if дата_окончания_списания:
            query += "дата_окончания_списания = %s AND "
            params_total.append(дата_окончания_списания)
        if договор_счет:
            query += "договор_счет ILIKE %s AND "
            params_total.append(f'%{договор_счет}%')
        if контрагент:
            query += "контрагент ILIKE %s AND "
            params_total.append(f'%{контрагент}%')
        if первичный_документ:
            query += "первичный_документ ILIKE %s AND "
            params_total.append(f'%{первичный_документ}%')
        if страна_производитель:
            query += "страна_производитель ILIKE %s AND "
            params_total.append(f'%{страна_производитель}%')
        if правообладатель:
            query += "правообладатель ILIKE %s AND "
            params_total.append(f'%{правообладатель}%')
        if срок_предоставления_права:
            query += "срок_предоставления_права ILIKE %s AND "
            params_total.append(f'%{срок_предоставления_права}%')
        if len(params_total) > 0:
            query = query[:-5]  
        else:
            return render_template('update_total.html', filtered_total=[])

        cur.execute(query, params_total)
        filtered_total = cur.fetchall()  
        cur.close()

        return render_template('update_total.html', filtered_total=filtered_total)

    return render_template('update_total.html', filtered_total=[])

editor_filtered_total = []
@app.route('/editor_submit_total', methods=['POST', 'GET'])
def editor_submit_total():
    global editor_filtered_total  
    if request.method == 'POST':
        статья_затрат = request.form.get('статья_затрат')
        наименование_ПО_БУ = request.form.get('наименование_ПО_БУ')
        наименование_ПО = request.form.get('наименование_ПО')
        краткое_наименование_ПО = request.form.get('краткое_наименование_ПО')
        код = request.form.get('код')
        филиал = request.form.get('филиал')
        счёт_затрат = request.form.get('счёт_затрат')
        вид_деятельности = request.form.get('вид_деятельности')
        срок_полезного_использования_мес = request.form.get('срок_полезного_использования_мес')
        дата_начала_списания = request.form.get('дата_начала_списания')
        дата_окончания_списания = request.form.get('дата_окончания_списания')
        договор_счет = request.form.get('договор_счет')
        контрагент = request.form.get('контрагент')
        первичный_документ = request.form.get('первичный_документ')
        страна_производитель = request.form.get('страна_производитель')
        правообладатель = request.form.get('правообладатель')
        срок_предоставления_права = request.form.get('срок_предоставления_права')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT статья_затрат, наименование_ПО_БУ, наименование_ПО, краткое_наименование_ПО, код, филиал, счет_затрат, вид_деятельности, срок_полезного_использования_мес, дата_начала_списания, дата_окончания_списания, договор_счет, контрагент, первичный_документ, страна_производитель, правообладатель, срок_предоставления_права FROM общая_информация WHERE "
        editor_params_total = []
        if статья_затрат:
            query += "статья_затрат = %s AND "
            editor_params_total.append(f'%{статья_затрат}%')
        if наименование_ПО_БУ:
            query += "наименование_ПО_БУ = %s AND "
            editor_params_total.append(f'%{наименование_ПО_БУ}%')
        if наименование_ПО:
            query += "наименование_ПО = %s AND "
            editor_params_total.append(f'%{наименование_ПО}%')
        if краткое_наименование_ПО: 
            query += "краткое_наименование_ПО = %s AND "
            editor_params_total.append(f'%{краткое_наименование_ПО}%')
        if код:
            query += "код = %s AND "
            editor_params_total.append(f'%{код}%')
        if филиал:
            query += "филиал ILIKE %s AND "
            editor_params_total.append(f'%{филиал}%')
        if счёт_затрат:
            query += "счёт_затрат = %s AND "
            editor_params_total.append(счёт_затрат)
        if вид_деятельности:
            query += "вид_деятельности ILIKE %s AND "
            editor_params_total.append(f'%{вид_деятельности}%')
        if срок_полезного_использования_мес:
            query += "срок_полезного_использования_мес = %s AND "
            editor_params_total.append(срок_полезного_использования_мес)
        if дата_начала_списания:
            query += "дата_начала_списания = %s AND "
            editor_params_total.append(дата_начала_списания)
        if дата_окончания_списания:
            query += "дата_окончания_списания = %s AND "
            editor_params_total.append(дата_окончания_списания)
        if договор_счет:
            query += "договор_счет ILIKE %s AND "
            editor_params_total.append(f'%{договор_счет}%')
        if контрагент:
            query += "контрагент ILIKE %s AND "
            editor_params_total.append(f'%{контрагент}%')
        if первичный_документ:
            query += "первичный_документ ILIKE %s AND "
            editor_params_total.append(f'%{первичный_документ}%')
        if страна_производитель:
            query += "страна_производитель ILIKE %s AND "
            editor_params_total.append(f'%{страна_производитель}%')
        if правообладатель:
            query += "правообладатель ILIKE %s AND "
            editor_params_total.append(f'%{правообладатель}%')
        if срок_предоставления_права:
            query += "срок_предоставления_права ILIKE %s AND "
            editor_params_total.append(f'%{срок_предоставления_права}%')
        if len(editor_params_total) > 0:
            query = query[:-5]  
        else:
            return render_template('editor_update_total.html', editor_filtered_total=[])

        cur.execute(query, editor_params_total)
        editor_filtered_total = cur.fetchall()  
        cur.close()

        return render_template('editor_update_total.html', editor_filtered_total=editor_filtered_total)

    return render_template('editor_update_total.html', editor_filtered_total=[])

@app.route('/export_total', methods=['POST', 'GET'])
def export_total():
    global filtered_total
    if filtered_total:
        df = pd.DataFrame(filtered_total)
        excel_file_path ='total.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404 
    
@app.route('/editor_export_total', methods=['POST', 'GET'])
def editor_export_total():
    global editor_filtered_total
    if editor_filtered_total:
        df = pd.DataFrame(editor_filtered_total)
        excel_file_path = 'total.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404

filtered_partner = []
@app.route('/submit_partner', methods=['POST', 'GET'])
def submit_partner():
    global filtered_partner
    if request.method == 'POST':
        наименование_контрагента = request.form.get('наименование_контрагента')
        краткое_наименование_предприятия = request.form.get('краткое_наименование_предприятия')
        полное_наименование_предприятия = request.form.get('полное_наименование_предприятия')
        ИНН = request.form.get('ИНН')
        КПП = request.form.get('КПП')
        юридический_адрес = request.form.get('юридический_адрес')
        фактический_адрес = request.form.get('фактический_адрес')
        ОКПО = request.form.get('ОКПО')
        ОГРН = request.form.get('ОГРН')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН FROM Контрагенты WHERE "
        partner_params = []
        if наименование_контрагента:
            query += "наименование_контрагента ILIKE %s AND "
            partner_params.append(f'%{наименование_контрагента}%')
        if краткое_наименование_предприятия:
            query += "краткое_наименование_предприятия ILIKE %s AND "
            partner_params.append(f'%{краткое_наименование_предприятия}%')
        if полное_наименование_предприятия:
            query += "полное_наименование_предприятия ILIKE %s AND "
            partner_params.append(f'%{полное_наименование_предприятия}%')
        if ИНН:
            query += "ИНН = %s AND "
            partner_params.append(ИНН)
        if КПП:
            query += "КПП = %s AND "
            partner_params.append(КПП)
        if юридический_адрес:
            query += "юридический_адрес ILIKE %s AND "
            partner_params.append(f'%{юридический_адрес}%')
        if фактический_адрес:
            query += "фактический_адрес ILIKE %s AND "
            partner_params.append(f'%{фактический_адрес}%')
        if ОКПО:
            query += "ОКПО = %s AND "
            partner_params.append(ОКПО)
        if ОГРН:
            query += "ОГРН = %s AND "
            partner_params.append(ОГРН)
        if len(partner_params) > 0:
            query = query[:-5] 
        else:
            return render_template('update_partner.html', filtered_partner=[])

        cur.execute(query, partner_params)
        filtered_partner = cur.fetchall()  
        cur.close()

        return render_template('update_partner.html', filtered_partner=filtered_partner)

    return render_template('update_partner.html', filtered_partner=[])

editor_filtered_partner = []
@app.route('/editor_submit_partner', methods=['POST', 'GET'])
def editor_submit_partner():
    global editor_filtered_partner
    if request.method == 'POST':
        наименование_контрагента = request.form.get('наименование_контрагента')
        краткое_наименование_предприятия = request.form.get('краткое_наименование_предприятия')
        полное_наименование_предприятия = request.form.get('полное_наименование_предприятия')
        ИНН = request.form.get('ИНН')
        КПП = request.form.get('КПП')
        юридический_адрес = request.form.get('юридический_адрес')
        фактический_адрес = request.form.get('фактический_адрес')
        ОКПО = request.form.get('ОКПО')
        ОГРН = request.form.get('ОГРН')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT наименование_контрагента, краткое_наименование_предприятия, полное_наименование_предприятия, ИНН, КПП, юридический_адрес, фактический_адрес, ОКПО, ОГРН FROM Контрагенты WHERE "
        editor_partner_params = []
        if наименование_контрагента:
            query += "наименование_контрагента ILIKE %s AND "
            editor_partner_params.append(f'%{наименование_контрагента}%')
        if краткое_наименование_предприятия:
            query += "краткое_наименование_предприятия ILIKE %s AND "
            editor_partner_params.append(f'%{краткое_наименование_предприятия}%')
        if полное_наименование_предприятия:
            query += "полное_наименование_предприятия ILIKE %s AND "
            editor_partner_params.append(f'%{полное_наименование_предприятия}%')
        if ИНН:
            query += "ИНН = %s AND "
            editor_partner_params.append(ИНН)
        if КПП:
            query += "КПП = %s AND "
            editor_partner_params.append(КПП)
        if юридический_адрес:
            query += "юридический_адрес ILIKE %s AND "
            editor_partner_params.append(f'%{юридический_адрес}%')
        if фактический_адрес:
            query += "фактический_адрес ILIKE %s AND "
            editor_partner_params.append(f'%{фактический_адрес}%')
        if ОКПО:
            query += "ОКПО = %s AND "
            editor_partner_params.append(ОКПО)
        if ОГРН:
            query += "ОГРН = %s AND "
            editor_partner_params.append(ОГРН)
        if len(editor_partner_params) > 0:
            query = query[:-5] 
        else:
            return render_template('editor_update_partner.html', editor_filtered_partner=[])

        cur.execute(query, editor_partner_params)
        editor_filtered_partner = cur.fetchall()  
        cur.close()

        return render_template('editor_update_partner.html', editor_filtered_partner=editor_filtered_partner)

    return render_template('editor_update_partner.html', editor_filtered_partner=[])

@app.route('/export_partner', methods=['POST', 'GET'])
def export_partner():
    global filtered_partner
    if filtered_partner:
        df = pd.DataFrame(filtered_partner)
        excel_file_path ='partner.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404 
    
@app.route('/editor_export_partner', methods=['POST', 'GET'])
def editor_export_partner():
    global editor_filtered_partner 
    if editor_filtered_partner:
        df = pd.DataFrame(editor_filtered_partner)
        excel_file_path = 'partner.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404

filtered_number = []
@app.route('/submit_number', methods=['POST', 'GET'])
def submit_number():
    global filtered_number
    if request.method == 'POST':
        номер_заявки = request.form.get('номер_заявки')
        наименование_ПО = request.form.get('наименование_ПО')
        тип_лицензии = request.form.get('тип_лицензии')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT номер_заявки, наименование_ПО, тип_лицензии FROM Учет_лицензий WHERE "
        number_params = []
        if номер_заявки:
            query += "номер_заявки = %s AND "
            number_params.append(номер_заявки)
        if наименование_ПО:
            query += "наименование_ПО ILIKE %s AND "
            number_params.append(f'%{наименование_ПО}%')
        if тип_лицензии:
            query += "тип_лицензии ILIKE %s AND "
            number_params.append(f'%{тип_лицензии}%')
        if len(number_params) > 0:
            query = query[:-5]
        else: 
            return render_template('update_number.html', filtered_number=[])
        cur.execute(query, number_params)
        filtered_number = cur.fetchall()
        return render_template('update_number.html', filtered_number=filtered_number)
    return render_template('update_number.html', filtered_number=[])

editor_filtered_number = []
@app.route('/editor_submit_number', methods=['POST', 'GET'])
def editor_submit_number():
    global editor_filtered_number
    if request.method == 'POST':
        номер_заявки = request.form.get('номер_заявки')
        наименование_ПО = request.form.get('наименование_ПО')
        тип_лицензии = request.form.get('тип_лицензии')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT номер_заявки, наименование_ПО, тип_лицензии FROM Учет_лицензий WHERE "
        editor_number_params = []
        if номер_заявки:
            query += "номер_заявки = %s AND "
            editor_number_params.append(номер_заявки)
        if наименование_ПО:
            query += "наименование_ПО ILIKE %s AND "
            editor_number_params.append(f'%{наименование_ПО}%')
        if тип_лицензии:
            query += "тип_лицензии ILIKE %s AND "
            editor_number_params.append(f'%{тип_лицензии}%')
        if len(editor_number_params) > 0:
            query = query[:-5]
        else: 
            return render_template('editor_update_number.html', editor_filtered_number=[])
        cur.execute(query, editor_number_params)
        editor_filtered_number = cur.fetchall()
        return render_template('editor_update_number.html', editor_filtered_number=editor_filtered_number)
    return render_template('editor_update_number.html', editor_filtered_number=[])

@app.route('/export_number', methods=['POST', 'GET'])
def export_number():
     global filtered_number
     if filtered_number:
         df = pd.DataFrame(filtered_number)
         excel_file_path ='number.xlsx'
         df.to_excel(excel_file_path, index=False)
         return send_file(excel_file_path, as_attachment=True)
     else:
         return "Нет данных для экспорта.", 404      
     
@app.route('/editor_export_number', methods=['POST', 'GET'])
def editor_export_number():
    global editor_filtered_number
    if editor_filtered_number:
        df = pd.DataFrame(editor_filtered_number)
        excel_file_path = 'number.xlsx'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404
    
filtered_software = []
@app.route('/submit_software', methods=['POST', 'GET'])
def submit_software():
    global filtered_software
    if request.method == 'POST':
        код_ПО = request.form.get('код_ПО')
        наименование_ПО = request.form.get('наименование')
        вендор = request.form.get('вендор')
        признак_ПО = request.form.get('признак_ПО')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT код_ПО, наименование_ПО, вендор, признак_ПО FROM Справочник_ПО WHERE "
        software_params = []
        if код_ПО:
            query += "код_ПО = %s AND "
            software_params.append(код_ПО)
        if наименование_ПО:
            query += "наименование_ПО ILIKE %s AND "
            software_params.append(f'%{наименование_ПО}%')
        if вендор:
            query += "вендор ILIKE %s AND "
            software_params.append(f'%{вендор}%')
        if признак_ПО:
            query += "признак_ПО ILIKE %s AND "
            software_params.append(f'%{признак_ПО}%')
        if len(software_params) > 0:
            query = query[:-5]
        else:
            return render_template('update_software.html', filtered_software=[])
        cur.execute(query, software_params)
        filtered_software = cur.fetchall()
        return render_template('update_software.html', filtered_software=filtered_software)
    return render_template('update_software.html', filtered_software=[])

editor_filtered_software = []
@app.route('/editor_submit_software', methods=['POST', 'GET'])
def editor_submit_software():
    global editor_filtered_software
    if request.method == 'POST':
        код_ПО = request.form.get('код_ПО')
        наименование_ПО = request.form.get('наименование')
        вендор = request.form.get('вендор')
        признак_ПО = request.form.get('признак_ПО')
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT код_ПО, наименование_ПО, вендор, признак_ПО FROM Справочник_ПО WHERE "
        editor_software_params = []
        if код_ПО:
            query += "код_ПО = %s AND "
            editor_software_params.append(код_ПО)
        if наименование_ПО:
            query += "наименование_ПО ILIKE %s AND "
            editor_software_params.append(f'%{наименование_ПО}%')
        if вендор:
            query += "вендор ILIKE %s AND "
            editor_software_params.append(f'%{вендор}%')
        if признак_ПО:
            query += "признак_ПО ILIKE %s AND "
            editor_software_params.append(f'%{признак_ПО}%')
        if len(editor_software_params) > 0:
            query = query[:-5]
        else:
            return render_template('editor_update_software.html', editor_filtered_software=[])
        cur.execute(query, editor_software_params)
        editor_filtered_software = cur.fetchall()
        return render_template('editor_update_software.html', editor_filtered_software=editor_filtered_software)
    return render_template('editor_update_software.html', editor_filtered_software=[])

@app.route('/export_software', methods=['POST', 'GET'])
def export_software():
     global filtered_software
     if filtered_software:
         df = pd.DataFrame(filtered_software)
         excel_file_path ='software.xlsx'
         df.to_excel(excel_file_path, index=False)
         return send_file(excel_file_path, as_attachment=True)
     else:
         return "Нет данных для экспорта.", 404    
     
@app.route('/editor_export_software', methods=['POST', 'GET'])
def editor_export_software():
    global editor_filtered_software
    if editor_filtered_software:
        df = pd.DataFrame(editor_filtered_software)
        excel_file_path = 'software.xlsx'
        df.to_Excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404

filtered = []
@app.route('/submit', methods=['POST', 'GET'])
def submit():
     global filtered 
     if request.method == 'POST':
         наименование_ПО = request.form.get('наименование_ПО')
         контрагент = request.form.get('контрагент')
         дата_начала_списания = request.form.get('дата_начала_списания')
         дата_окончания_списания = request.form.get('дата_окончания_списания')
         признак_ПО = request.form.get('признак_ПО')
         страна_производитель = request.form.get('страна_производитель')
         код = request.form.get('код')
         cur = conn2.cursor(cursor_factory = psycopg2.extras.DictCursor)
         lic_params = []
         query = "SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, признак_ПО, страна_производитель, код FROM лицензии WHERE "
         if наименование_ПО:
             query += "наименование_ПО ILIKE %s AND "
             lic_params.append(f'%{наименование_ПО}%')
         if контрагент:
             query += "контрагент ILIKE %s AND "
             lic_params.append(f'%{контрагент}%')
         if дата_начала_списания:
             query += "дата_начала_списания = %s AND "
             lic_params.append(дата_начала_списания)
         if дата_окончания_списания:
             query += "дата_окончания_списания = %s AND "
             lic_params.append(дата_окончания_списания)
         if признак_ПО:
             query += "признак_ПО ILIKE %s AND "
             lic_params.append(f'%{признак_ПО}%')
         if страна_производитель:
             query += "страна_производитель ILIKE %s AND "
             lic_params.append(f'%{страна_производитель}%')
         if код:
             query += "код = %s AND "
             lic_params.append(код)
         if len(lic_params) > 0:
             query = query[:-5]
         else:
             return render_template('update_lic.html', filtered=[])
         cur.execute(query, lic_params)
         filtered = cur.fetchall()
         return render_template('update_lic.html', filtered=filtered)
     return render_template('update_lic.html', filtered=[])
 
editor_filtered = []
@app.route('/editor_submit', methods=['POST', 'GET'])
def editor_submit():
    global editor_filtered 
    if request.method == 'POST':
        наименование_ПО = request.form.get('наименование_ПО')
        контрагент = request.form.get('контрагент')
        дата_начала_списания = request.form.get('дата_начала_списания')
        дата_окончания_списания = request.form.get('дата_окончания_списания')
        признак_ПО = request.form.get('признак_ПО')
        страна_производитель = request.form.get('страна_производитель')
        код = request.form.get('код')
        cur = conn2.cursor(cursor_factory = psycopg2.extras.DictCursor)
        editor_lic_params = []
        query = "SELECT наименование_ПО, контрагент, дата_начала_списания, дата_окончания_списания, признак_ПО, страна_производитель, код FROM лицензии WHERE "
        if наименование_ПО:
            query += "наименование_ПО ILIKE %s AND "
            editor_lic_params.append(f'%{наименование_ПО}%')
        if контрагент:
            query += "контрагент ILIKE %s AND "
            editor_lic_params.append(f'%{контрагент}%')
        if дата_начала_списания:
            query += "дата_начала_списания = %s AND "
            editor_lic_params.append(дата_начала_списания)
        if дата_окончания_списания:
            query += "дата_окончания_списания = %s AND "
            editor_lic_params.append(дата_окончания_списания)
        if признак_ПО:
            query += "признак_ПО ILIKE %s AND "
            editor_lic_params.append(f'%{признак_ПО}%')
        if страна_производитель:
            query += "страна_производитель ILIKE %s AND "
            editor_lic_params.append(f'%{страна_производитель}%')
        if код:
            query += "код = %s AND "
            editor_lic_params.append(код)
        if len(editor_lic_params) > 0:
            query = query[:-5]
        else:
            return render_template('editor_update_lic.html', editor_filtered=[])
        cur.execute(query, editor_lic_params)
        editor_filtered = cur.fetchall()
        return render_template('editor_update_lic.html', editor_filtered=editor_filtered)
    return render_template('editor_update_lic.html', editor_filtered=[])
 
@app.route('/export', methods=['POST', 'GET'])
def export():
     global filtered
     if filtered:
         df = pd.DataFrame(filtered)
         excel_file_path ='filtered_lic.xlsx'
         df.to_excel(excel_file_path, index=False)
         return send_file(excel_file_path, as_attachment=True)
     else:
         return "Нет данных для экспорта.", 404 
     
@app.route('/editor_export', methods=['POST', 'GET'])
def editor_export():
    global editor_filtered
    if editor_filtered:
        df = pd.DataFrame(editor_filtered)
        excel_file_path = 'filtered_lic.html'
        df.to_excel(excel_file_path, index=False)
        return send_file(excel_file_path, as_attachment=True)
    else:
        return "Нет данных для экспорта.", 404

filtered_install = []    
@app.route('/submit_install', methods=['POST', 'GET'])
def submit_install():
    global filtered_install
    if request.method == 'POST':
        код_установки  = request.form.get('код_установки')
        наименование_ПО = request.form.get('наименование_ПО')
        тип_лицензии = request.form.get('тип_лицензии')
        дата_установки_ПО = request.form.get('дата_установки_ПО')
        install_params = []
        cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        query = "SELECT код_установки, наименование_ПО, тип_лицензии, дата_установки_ПО FROM Установка_ПО WHERE "
        if код_установки:
            query += "код_установки = %s AND "
            install_params.append(код_установки)
        if наименование_ПО:
            query += "наиенование_ПО ILIKE %s AND "
            install_params.append(f'%{наименование_ПО}%')
        if тип_лицензии:
            query += "тип_лицензии ILIKE %s AND "
            install_params.append(f'%{тип_лицензии}%')
        if дата_установки_ПО:
            query += "дата_установки_ПО = %s AND "
            install_params.append(дата_установки_ПО)
        if len(install_params) > 0:
            query = query[:-5]
        else:
            return render_template('update_install.html', filtered_install=[])
        cur.execute(query, install_params)
        filtered_install = cur.fetchall()
        return render_template('update_install.html', filtered_install=filtered_install)
    return render_template('update_install.html', filtered_install=[])

@app.route('/export_install', methods=['POST', 'GET'])
def export_install():
     global filtered_install
     if filtered_install:
         df = pd.DataFrame(filtered_install)
         excel_file_path ='filtered_install.xlsx'
         df.to_excel(excel_file_path, index=False)
         return send_file(excel_file_path, as_attachment=True)
     else:
         return "Нет данных для экспорта.", 404 
     
@app.route('/user_change_password', methods = ['POST', 'GET'])
def user_change_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
        account = cur.fetchone()
        if request.method == 'POST':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            if not old_password or not new_password:
                flash('Пожалуйста, заполните форму!')
            else:
                if account:
                        if not check_password_hash(account['password'], old_password):
                            flash('Неверный старый пароль!')
                            return redirect(url_for('user_change_password'))
                        if old_password == new_password:
                            flash('Пароли совпадают!')
                            return redirect(url_for('user_change_password'))
                        id=account['id']
                        _hashed_password = generate_password_hash(new_password)
                        cur.execute("""UPDATE users
                                    SET password=%s
                                    WHERE id=%s
                                    """, (_hashed_password, id))
                        flash('Запись успешно обновлена!')
                        conn.commit()
                        return redirect(url_for('profile'))
                else:
                    flash('Введены некорректные данные!')
                    return redirect(url_for('user_change_password'))
        return render_template('profile_change_password.html', title="Поменять пароль")
    
@app.route('/support_change_password', methods=['POST', 'GET'])
def support_change_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
        account = cur.fetchone()
        if request.method == 'POST':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            if not old_password or not new_password:
                flash('Пожалуйста, заполните форму!')
            else:
                if account:
                    if not check_password_hash(account['password'], old_password):
                        flash('Неверный старый пароль!')
                        return redirect(url_for('support_change_password'))
                    if old_password == new_password:
                        flash('Пароли совпадают!')
                        return redirect(url_for('support_change_password'))
                    id=account['id']
                    _hashed_password = generate_password_hash(new_password)
                    cur.execute("""UPDATE users
                                SET password=%s
                                WHERE id=%s""", (_hashed_password, id))
                    flash('Запись успешно обновлена!')
                    conn.commit()
                    return redirect(url_for('support_profile'))
                else:
                    flash('Введены некорректные данные!')
                    return redirect(url_for('support_change_password'))
        return render_template('support_change_password.html', title='Поменять пароль')
    
@app.route('/editor_change_password', methods=['POST', 'GET'])
def editor_change_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
        account = cur.fetchone()
        if request.method == 'POST':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            if not old_password or not new_password:
                flash('Пожалуйста, заполните форму!')
            else:
                if account:
                    if not check_password_hash(account['password'], old_password):
                        flash('Неверный старый пароль!')
                        return redirect(url_for('editor_change_password'))
                    if old_password == new_password:
                        flash('Пароли совпадают!')
                        return redirect(url_for('editor_change_password'))
                    id=account['id']
                    _hashed_password = generate_password_hash(new_password)
                    cur.execute("""UPDATE users
                                SET password=%s
                                WHERE id=%s""", (_hashed_password, id))
                    flash('Запись успешно обновлена!')
                    conn.commit()
                    return redirect(url_for('editor_profile'))
                else:
                    flash('Введены некорректные данные!')
                    return redirect(url_for('editor_change_password'))
        return render_template('editor_change_password.html', title='Поменять пароль')
    

@app.route('/user_change_email', methods=['POST', 'GET'])
def user_change_email():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            new_email = request.form.get('new_email')
            if not new_email:
                flash("Пожалуйста, заполните форму!")
            else:
                cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
                account = cur.fetchone()
                if account:
                    if new_email == session['email']:
                        flash('Новый адрес электронной почты совпадает с предыдущим!')
                        return redirect(url_for('user_change_email'))
                    id = account['id']
                    cur.execute("""UPDATE users
                                SET email=%s
                                WHERE id=%s""", (new_email, id))
                    flash('Запись успешно обновлена!')
                    conn.commit()
                    session['email'] = new_email  
                    return redirect(url_for('profile'))
                else:
                    flash('Введен некорректный адрес электронной почты!')
                    return redirect(url_for('user_change_email'))
        return render_template('change_email.html', title='Поменять адрес электронной почты')

@app.route('/support_change_email', methods=['POST', 'GET'])
def support_change_email():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            new_email = request.form.get('new_email')
            if not new_email:
                flash('Пожалуйста, заполните форму!')
            else:
                cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
                account = cur.fetchone()
                if account:
                    if new_email == session['email']:
                        flash('Новый адрес электронной почты совпадает с предыдущим!')
                        return redirect(url_for('support_change_email'))
                    id = account['id']
                    cur.execute("""UPDATE users
                                SET email=%s
                                WHERE id=%s""", (new_email, id))
                    flash('Запись успешно обновлена!')
                    conn.commit()
                    session['email'] = new_email
                    return redirect(url_for('support_profile'))
                else:
                    flash('Введен некорректный адрес электронной почты!')
                    return redirect(url_for('support_change_email'))
        return render_template('support_change_email.html', title='Поменять адрес электронной почты')

@app.route('/editor_change_email', methods=['POST', 'GET'])
def editor_change_email():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    else:
        if request.method == 'POST':
            new_email = request.form.get('new_email')
            if not new_email:
                flash('Пожалуйста, заполните форму!')
            else:
                cur.execute('SELECT * FROM users WHERE username=%s', (session['username'],))
                account = cur.fetchone()
                if account:
                    if new_email == session['email']:
                        flash('Новый адрес электронной почты совпадает с предыдущим!')
                        return redirect(url_for('editor_change_email'))
                    id = account['id']
                    cur.execute("""UPDATE users
                                SET email=%s
                                WHERE id=%s""", (new_email, id))
                    flash('Запись успешно обновлена!')
                    conn.commit()
                    session['email'] = new_email
                    return redirect(url_for('editor_profile'))
                else:
                    flash('Введен некорректный адрес электронной почты!')
                    return redirect(url_for('editor_change_email'))
        return render_template('editor_change_email.html', title='Поменять адрес электронной почты')

@app.route('/profile_reset_password', methods=['POST', 'GET'])
def profile_reset_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        email = request.form.get('email')
        user_otp = request.form.get('otp')
        current_email = session.get('email')
        if email != current_email:
            flash('Введенный адрес электронной почты не совпадает с предыдущим!')
            return render_template('profile_reset_password.html')
        cur.execute('SELECT * FROM users WHERE email=%s', (session['email'],))
        user = cur.fetchone()
        if not user:
            flash('Неправильный адрес электронной почты! Пожалуйста, попробуйте еще раз.')
            return render_template('profile_reset_password.html') 
        if user_otp:
            otp = session.get('otp')
            if otp == int(user_otp):
                return redirect(url_for('profile_change_password', email=email))
            else:
                flash('Неверный код подтверждения!')
                return render_template('profile_reset_password.html', email=email)
        else:
            otp = randint(000000, 999999)
            send_email(email, otp)
            session['otp'] = otp
            flash('Код подтверждения отправлен вам на адрес электронной почты!')
            return render_template('profile_reset_password.html', email=email)
    return render_template('profile_reset_password.html', title='Пользователь забыл пароль')

@app.route('/profile_support_reset_password', methods=['POST', 'GET'])
def profile_support_reset_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        email = request.form.get('email')
        user_otp = request.form.get('otp')
        current_email = session.get('email')
        if email != current_email:
            flash('Введенный адрес электронной почты не совпадает с предыдущим!')
            return render_template('profile_support_reset_password.html')
        cur.execute('SELECT * FROM users WHERE email=%s', (session['email'],))
        user = cur.fetchone()
        if not user:
            flash('Неправильный адрес электронной почты! Пожалуйста, попробуйте еще раз.')
            return render_template('profile_support_reset_password.html') 
        if user_otp:
            otp = session.get('otp')
            if otp == int(user_otp):
                return redirect(url_for('profile_support_change_password', email=email))
            else:
                flash('Неверный код подтверждения!')
                return render_template('profile_support_reset_password.html', email=email)
        else:
            otp = randint(000000, 999999)
            send_email(email, otp)
            session['otp'] = otp
            flash('Код подтверждения отправлен вам на адрес электронной почты!')
            return render_template('profile_support_reset_password.html', email=email)
    return render_template('profile_support_reset_password.html', title='Пользователь забыл пароль')

@app.route('/profile_editor_reset_password', methods=['POST', 'GET'])
def profile_editor_reset_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        email = request.form.get('email')
        user_otp = request.form.get('otp')
        current_email = session.get('email')
        if email != current_email:
            flash('Введенный адрес электронной почты не совпадает с предыдущим!')
            return render_template('profile_editor_reset_password.html')
        cur.execute('SELECT * FROM users WHERE email=%s', (session['email'],))
        user = cur.fetchone()
        if not user:
            flash('Неправильный адрес электронной почты! Пожалуйста, попробуйте еще раз. ')
            return render_template('profile_editor_reset_password.html')
        if user_otp:
            otp = session.get('otp')
            if otp == int(user_otp):
                return redirect(url_for('profile_editor_change_password', email=email))
            else:
                flash('Неверный код подтверждения!')
                return render_template('profile_editor_reset_password.html', email=email)
        else:
            otp = randint(000000, 999999)
            send_email(email, otp)
            session['otp'] = otp
            flash('Код подтверждения отправлен вам на адрес электронной почты!')
            return render_template('profile_editor_reset_password.html', email=email)
    return render_template('profile_editor_reset_password.html', title='Пользователь забыл пароль ')

@app.route('/reset_email', methods=['POST', 'GET'])
def reset_email():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    if request.method == 'POST':
        phone_number = request.form.get('phone_number')
        cur.execute('SELECT * FROM users WHERE phone_number=%s', (phone_number,))
        user = cur.fetchone()
        
        if not user:
            flash('Неправильный номер телефона! Пожалуйста, попробуйте еще раз.')
            return render_template('reset_email.html')
        else:
            return redirect(url_for('change_password_email', phone_number=phone_number))
    return render_template('reset_email.html', title='Пользователь забыл адрес электронной почты')
    

@app.route('/reset_password', methods=['POST', 'GET'])
def reset_password():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == 'POST':
        email = request.form.get('email')
        user_otp = request.form.get('otp')

        cur.execute('SELECT * FROM users WHERE email=%s', (email,))
        user = cur.fetchone()

        if not user:
            flash('Неправильный адрес электронной почты! Пожалуйста, попробуйте еще раз.')
            return render_template('reset_password.html')

        if user_otp:
            otp = session.get('otp')
            if otp == int(user_otp):
                return redirect(url_for('change_password', email=email))
            else:
                flash('Неверный код подтверждения!')
                return render_template('reset_password.html', email=email)
        else:
            otp = randint(000000, 999999)
            send_email(email, otp)
            session['otp'] = otp
            flash('Код подтверждения отправлен вам на адрес электронной почты!')
            return render_template('reset_password.html', email=email)

    return render_template('reset_password.html', title='Пользователь забыл пароль')

def send_email(email, otp):
    msg = MIMEText(str(otp))
    msg['Subject'] = 'OTP'
    msg['From'] = 'Analstasia.Sholokhova@yandex.ru'
    msg['To'] = email

    with smtplib.SMTP('smtp.yandex.ru', 587) as smtp:
        smtp.starttls()
        smtp.login('Analstasia.Sholokhova@yandex.ru', 'xyanueciccoxachh')
        smtp.send_message(msg)
        
@app.route('/change_password_email/<phone_number>', methods=['POST', 'GET'])
def change_password_email(phone_number):
    cur= conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT password FROM users WHERE phone_number=%s', (phone_number,))
    user = cur.fetchone()
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        if check_password_hash(user['password'], new_password):
            flash('Новый пароль совпадает с текущим!')
            return redirect(url_for('change_password_email', phone_number=phone_number))
        _hashed_password = generate_password_hash(new_password)
        cur.execute("""UPDATE users
                    SET password=%s
                    WHERE phone_number=%s""", (_hashed_password, phone_number))
        flash('Пароль успешно обновлен!')
        conn.commit()
        return redirect(url_for('login'))
    return render_template('change_password_email.html', phone_number=phone_number, title='Изменить пароль')

@app.route('/change_password/<email>', methods=['POST', 'GET'])
def change_password(email):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT password FROM users WHERE email=%s', (email,))
    user = cur.fetchone()
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        if check_password_hash(user['password'], new_password):
            flash('Новый пароль совпадает с текущим паролем!')
            return redirect(url_for('change_password', email=email))
        _hashed_password = generate_password_hash(new_password)
        cur.execute("""UPDATE users
                    SET password=%s
                    WHERE email=%s""", (_hashed_password, email))
        flash('Пароль успешно обновлен!')
        conn.commit()
        return redirect(url_for('login'))
    return render_template('change_password.html', email=email, title='Изменить пароль')

@app.route('/profile_change_password/<email>', methods=['POST', 'GET'])
def profile_change_password(email):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT password FROM users WHERE email=%s", (email,))
    user = cur.fetchone()
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        if check_password_hash(user['password'], new_password):
            flash('Новый пароль совпадает с текущим паролем!')
            return redirect(url_for('profile_change_password', email=email))
        _hashed_password = generate_password_hash(new_password)
        cur.execute("""UPDATE users
                    SET password=%s
                    WHERE email=%s""", (_hashed_password, email))
        flash('Пароль успешно обновлен!')
        conn.commit()
        return redirect(url_for('profile'))
    return render_template('success_change_password.html', email=email, title='Изменить пароль')

@app.route('/profile_support_change_password/<email>', methods=['POST', 'GET'])
def profile_support_change_password(email):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT password FROM users WHERE email=%s", (email,))
    user = cur.fetchone()
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        if check_password_hash(user['password'], new_password):
            flash('Новый пароль совпадает с текущим паролем!')
            return redirect(url_for('profile_support_change_password', email=email))
        _hashed_password = generate_password_hash(new_password)
        cur.execute("""UPDATE users
                    SET password=%s
                    WHERE email=%s""", (_hashed_password, email))
        flash('Пароль успешно обновлен!')
        conn.commit()
        return redirect(url_for('support_profile'))
    return render_template('success_support_change_password.html', email=email, title='Изменить пароль')

@app.route('/profile_editor_change_password/<email>', methods=['POST', 'GET'])
def profile_editor_change_password(email):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT( password FROM users WHERE email=%s', (email,))
    user = cur.fetchone()
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        if check_password_hash(user['password'], new_password):
            flash('Новый пароль совпадает с текущим паролем!')
            return redirect(url_for('profile_editor_change_password', email=email))
        _hashed_password = generate_password_hash(new_password)
        cur.execute("""UPDATE users
                    SET password=%s
                    WHERE email=%s""", (_hashed_password, email))
        flash('Пароль успешно обновлен!')
        conn.commit()
        return redirect(url_for('editor_profile'))
    return render_template('success_editor_change_password.html', email=email, title='Изменить пароль')


@app.route('/page/<software>')
def show_installation_details(software):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT наименование_ПО, ФИО, отдел, ip_адрес, наименование_машины, дата_установки_ПО, тип_лицензии FROM Установка_ПО WHERE наименование_ПО=%s AND чекбокс=True", (software,))
    list_installation = cur.fetchall()
    return render_template('software_description.html', list_installation=list_installation, software=software)

@app.route('/support_page/<software>')
def support_show_installation_details(software):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("""
            SELECT наименование_ПО, ФИО, отдел, ip_адрес, наименование_машины, дата_установки_ПО, тип_лицензии
            FROM Установка_ПО 
            WHERE наименование_ПО = %s AND чекбокс=True
        """, (software,))
    list_installation = cur.fetchall()
    return render_template('support_software_description.html', list_installation=list_installation, software=software)


@app.route('/download/soft_report/excel/<software>')
def download_soft_report(software):
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT код_установки, ФИО, отдел, ip_адрес, наименование_машины, дата_установки_ПО, тип_лицензии FROM Установка_ПО WHERE наименование_ПО=%s AND чекбокс=True', (software,))
    result = cur.fetchall()
    output = io.BytesIO()
    workbook = xlwt.Workbook()
    sh = workbook.add_sheet('Отчет')
    sh.write(0, 0, 'Код установки')
    sh.write(0, 1, 'ФИО')
    sh.write(0, 2, 'Отдел')
    sh.write(0, 3, 'IP адрес')
    sh.write(0, 4, 'Наименование машины')
    sh.write(0, 5, 'Дата установки ПО')
    sh.write(0, 6, 'Тип лицензии')
    idx = 0
    for row in result:
        sh.write(idx+1, 0, str(row['код_установки']))
        sh.write(idx+1, 1, row['ФИО'])
        sh.write(idx+1, 2, row['отдел'])
        sh.write(idx+1, 3, row['ip_адрес'])
        sh.write(idx+1, 4, row['наименование_машины'])
        sh.write(idx+1, 5, row['дата_установки_ПО'])
        sh.write(idx+1, 6, row['тип_лицензии'])
        idx += 1
    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=soft_report.xls"})

@app.route('/upload_data', methods=['POST', 'GET'])
def upload_data():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    if len(row) == 15: 
                        data.append(row)
                    else:
                        flash('Неправильный формат файла')
                        return redirect(url_for('home'))
                for row in data:
                    cur.execute(
                        "INSERT INTO лицензии (наименование_ПО, вендор, начало_действия_лицензии, окончание_действия_лицензии, счёт_списания, стоимость_за_единицу, итоговая_стоимость, заказчик_ПО, признак_ПО, количество_ПО, срок_действия_лицензии, оплачено, остаток, примечание, номер_пп) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, (to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW())::interval, %s, %s, %s, %s)",
                        (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14])
                    )
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('home'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('home'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('home'))
        return render_template('home.html')
    else:
        return render_template('home.html')
    
@app.route('/editor_upload_data', methods=['POST', 'GET'])
def editor_upload_data():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    if len(row) == 15: 
                        data.append(row)
                    else:
                        flash('Неправильный формат файла')
                        return redirect(url_for('home_editor'))
                for row in data:
                    cur.execute(
                        "INSERT INTO лицензии (номер_пп, наименование_ПО, вендор, начало_действия_лицензии, окончание_действия_лицензии, счёт_списания, стоимость_за_единицу, итоговая_стоимость, заказчик_ПО, признак_ПО, количество_ПО, срок_действия_лицензии, оплачено, остаток, примечание) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, (to_timestamp(CAST(%s AS text), 'YYYY-MM-DD') - NOW())::interval, %s, %s, %s)",
                        (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14])
                    )
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('home_editor'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('home_editor'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('home_editor'))
        return render_template('home_editor.html')
    else:
        return render_template('home_editor.html')

@app.route('/upload_licence', methods=['POST', 'GET'])
def upload_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):#костыль
                    data.append(row)
                for row in data:
                    cur.execute('INSERT INTO Справочник_лицензий (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание) VALUES(%s,%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4], row[5]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('licence_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('licence_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('licence_list'))
        return render_template('licence.html')
    else:
        return render_template('licence.html')
    
@app.route('/editor_upload_licence', methods=['POST', 'GET'])
def editor_upload_licence():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    cur.execute('INSERT INTO Справочник_лицензий (код_лицензии, наименование_лицензии, тип_лицензии, счёт_списания, версия_лицензии, примечание) VALUES(%s,%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4], row[5]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_licence_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_licence_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_licence_list'))
        return render_template('editor_licence.html')
    else:
        return render_template('editor_licence.html')
    
@app.route("/upload_software", methods=['POST', 'GET'])
def upload_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-2):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute("INSERT INTO Справочник_ПО (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)", (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('software_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('software_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('software_list'))
        return render_template('software.html')
    else:
        return render_template('software.html')
    
@app.route('/editor_upload_software', methods=['POST', 'GET'])
def editor_upload_software():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-2):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute("INSERT INTO Справочник_ПО (код_ПО, наименование_ПО, описание_ПО, ссылка_на_сайт_ПО, вендор, стоимость_за_единицу, признак_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)", (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_software_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_software_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_software_list'))
        return render_template('editor_software.html')
    else:
        return render_template('editor_software.html')
    
@app.route('/upload_customer', methods=['POST', 'GET'])
def upload_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Справочник_заказчиков_ПО (код_заказчика, заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание) VALUES(%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('customer_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('customer_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('customer_list'))
        return render_template('customer.html')
    else:
        return render_template('customer.html')
    
@app.route('/editor_upload_customer', methods=['POST', 'GET'])
def editor_upload_customer():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Справочник_заказчиков_ПО (код_заказчика, заказчик_ПО, описание_заказчика, ссылка_на_сайт_заказчика, примечание) VALUES(%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_customer_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_customer_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_customer_list'))
        return render_template('editor_customer.html')
    else:
        return render_template('editor_customer.html')
    
@app.route('/upload_vendor', methods=['POST', 'GET'])
def upload_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Справочник_производителей_ПО (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание) VALUES(%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('vendor_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('vendor_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('vendor_list'))
        return render_template('vendor.html')
    else:
        return render_template('vendor.html')
    
@app.route('/editor_upload_vendor', methods=['POST', 'GET'])
def editor_upload_vendor():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Справочник_производителей_ПО (код_производителя, производитель, описание_производителя, ссылка_на_сайт_производителя, примечание) VALUES(%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_vendor_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_vendor_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_vendor_list'))
        return render_template('editor_vendor.html')
    else:
        return render_template('editor_vendor.html')
    
@app.route('/upload_installation', methods=['POST', 'GET'])
def upload_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, ФИО, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('install_software'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('install_software'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('install_software'))
        return render_template('install.html')
    else:
        return render_template('install.html')
    
@app.route('/support_upload_installation')
def support_upload_installation():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Установка_ПО (код_установки, наименование_ПО, тип_лицензии, ФИО, ip_адрес, наименование_машины, чекбокс, общее_количество, число_установленных_лицензий, дата_установки_ПО, чекбокс_условно_бесплатное_ПО, примечание) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('support_install_software'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('support_install_software'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('support_install_software'))
        return render_template('support_install.html')
    else:
        return render_template('support_install.html')
    
@app.route('/upload_number', methods=['POST', 'GET'])
def upload_number():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Учет_лицензий (номер_заявки, наименование_ПО, тип_лицензии, количество_лицензий_ПО, примечание) VALUES(%s,%s,%s,%s,%s)', (row[0], row[1], row[2], row[3], row[4]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('number_licences'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('number_licences'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('number_licences'))
        return render_template('number.html')
    else:
        return render_template('number.html')

@app.route('/upload_partner', methods=['POST', 'GET'])
def upload_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Контрагенты (код_контрагента, наименование_контрагента, договор, примечание) VALUES(%s,%s,%s,%s)', (row[0], row[1], row[2], row[3]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('partners_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('partners_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('partners_list'))
        return render_template('partner.html')
    else:
        return render_template('partner.html')

@app.route('/editor_upload_partner', methods=['POST', 'GET'])
def editor_upload_partner():
    cur = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method=='POST':
        file = request.files['file']
        if file:
            if file.filename.endswith("xlsx"):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True, min_row=2, max_row=sheet.max_row-1):
                    data.append(row)
                for row in data:
                    print(row)
                    cur.execute('INSERT INTO Контрагенты (код_контрагента, наименование_контрагента, договор, примечание) VALUES(%s,%s,%s,%s)', (row[0], row[1], row[2], row[3]))
                conn2.commit()
                flash('Файл успешно загружен!')
                return redirect(url_for('editor_partners_list'))
            else:
                flash('Неверный тип файла')
                return redirect(url_for('editor_partners_list'))
        else:
            flash('Файл не выбран!')
            return redirect(url_for('editor_partners_list'))
        return render_template('editor_partner.html')
    else:
        return render_template('editor_partner.html')
    
if __name__ == "__main__":
    serve(app, host="127.0.0.1", port=5000)


